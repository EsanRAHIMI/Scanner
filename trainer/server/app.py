import asyncio
import json
import logging
import os
import random
import re
import shutil
import subprocess
import tempfile
import threading
import time
import uuid
import base64
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal
from urllib.parse import urlencode
from urllib.request import Request

from typing_extensions import TypedDict

import yaml
from fastapi import Depends, FastAPI, File, HTTPException, Query, Request as FastAPIRequest, Response, UploadFile
from fastapi.responses import FileResponse
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from PIL import Image

import io

from fastapi.middleware.cors import CORSMiddleware

import jwt
from email_validator import EmailNotValidError, validate_email
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pymongo import InsertOne, UpdateOne

CLASSES_COLLECTION = "trainer_classes"
QUEUE_COLLECTION = "trainer_queue"
PRODUCT_FIELD_OPTIONS_COLLECTION = "product_field_options"
PRODUCT_FIELD_OPTIONS_DOC_ID = "selectable_fields"
CONTENT_CALENDAR_FIELD_OPTIONS_COLLECTION = "content_calendar_field_options"
CONTENT_CALENDAR_FIELD_OPTIONS_DOC_ID = "selectable_fields"
CONTENT_CALENDAR_SELECTABLE_FIELDS = (
  "Target Audience",
  "Content Pillar",
  "Tone of Voice",
  "Status",
  "# Hashtag",
  "Format",
)
CONTENT_CALENDAR_MULTI_VALUE_FIELDS = {"Target Audience", "# Hashtag"}
DEFAULT_CALENDAR_FIELD_OPTIONS: dict[str, list[str]] = {
  "Status": ["Published", "Scheduled", "In Progress", "Drafts"],
}
PRODUCT_IMPORT_BATCHES_COLLECTION = "product_import_batches"
PRODUCT_IMPORT_ROWS_COLLECTION = "product_import_rows"
PRODUCT_SELECTABLE_FIELDS = ("Category", "Space", "Color", "Material")
PRODUCT_IMPORT_CANONICAL_COLUMNS = [
  "Row",
  "Image1",
  "Image",
  "DAM",
  "Video",
  "Price",
  "URL",
  "Colecction Name",
  "Colecction Code",
  "Factory Code",
  "Variant Number",
  "Category",
  "Space",
  "Color",
  "Material",
  "DIMENSION (mm)",
  "Note",
  "Details",
  "CODE NUMBER",
  "L000",
  "Num",
  "Main",
  "Content Calendar",
]
PRODUCT_IMPORT_HEADER_ALIASES: dict[str, str] = {
  "image": "Image",
  "images": "Image",
  "photo": "Image",
  "photos": "Image",
  "picture": "Image",
  "dam": "DAM",
  "dam link": "DAM",
  "asset": "DAM",
  "asset url": "DAM",
  "video": "Video",
  "video url": "Video",
  "price": "Price",
  "قیمت": "Price",
  "amount": "Price",
  "cost": "Price",
  "url": "URL",
  "link": "URL",
  "links": "URL",
  "collection": "Colecction Name",
  "collection name": "Colecction Name",
  "colecction name": "Colecction Name",
  "product name": "Colecction Name",
  "name": "Colecction Name",
  "نام": "Colecction Name",
  "collection code": "Colecction Code",
  "colecction code": "Colecction Code",
  "code": "Colecction Code",
  "factory code": "Factory Code",
  "factory": "Factory Code",
  "factorycode": "Factory Code",
  "variant": "Variant Number",
  "variant number": "Variant Number",
  "variant no": "Variant Number",
  "variant #": "Variant Number",
  "category": "Category",
  "space": "Space",
  "room": "Space",
  "color": "Color",
  "colour": "Color",
  "material": "Material",
  "dimension": "DIMENSION (mm)",
  "dimensions": "DIMENSION (mm)",
  "dimension cm": "DIMENSION (mm)",
  "dimension mm": "DIMENSION (mm)",
  "size": "DIMENSION (mm)",
  "note": "Note",
  "notes": "Note",
  "description": "Note",
  "detail": "Details",
  "details": "Details",
  "code number": "CODE NUMBER",
  "code no": "CODE NUMBER",
  "l000": "L000",
  "num": "Num",
  "number": "Num",
  "main": "Main",
  "main variant": "Main",
  "content calendar": "Content Calendar",
}
DEFAULT_PRODUCT_FIELD_OPTIONS: dict[str, list[str]] = {
  "Category": [
    "Chandeliers",
    "Pendant",
    "Cascade Light",
    "Floor Lamps",
    "Long Chandeliers",
    "Ring Chandeliers",
    "Wall Light",
    "Table Lamps",
    "Accessories",
    "Sofa & Seating",
    "Table",
    "Wall Decoration",
  ],
  "Space": [
    "Corner",
    "Corridor",
    "Entrance",
    "Staircase",
    "Living Room",
    "Dining Room",
    "Bedroom",
    "Kitchen",
    "Commercial",
    "Bathroom",
  ],
  "Color": ["Transparent", "Chrome", "White", "Black", "Bronze", "Blue", "Gold", "Pink"],
  "Material": ["Stone", "Fabric", "Metal", "Glass", "Wood"],
}


def _normalize_product_field_options_payload(raw_options: Any) -> dict[str, list[str]]:
  if not isinstance(raw_options, dict):
    raise HTTPException(status_code=400, detail="INVALID_OPTIONS_PAYLOAD")

  out: dict[str, list[str]] = {}
  for field_name in PRODUCT_SELECTABLE_FIELDS:
    raw_values = raw_options.get(field_name, [])
    if not isinstance(raw_values, list):
      raise HTTPException(status_code=400, detail=f"INVALID_OPTIONS_FOR_{field_name.upper()}")

    seen: set[str] = set()
    cleaned: list[str] = []
    for item in raw_values:
      if not isinstance(item, str):
        continue
      value = item.strip()
      key = value.casefold()
      if not value or key in seen:
        continue
      seen.add(key)
      cleaned.append(value)

    out[field_name] = cleaned

  return out


async def _derive_product_field_options(db: Any) -> dict[str, list[str]]:
  options: dict[str, set[str]] = {field_name: set() for field_name in PRODUCT_SELECTABLE_FIELDS}
  projection = {f"fields.{field_name}": 1 for field_name in PRODUCT_SELECTABLE_FIELDS}
  cursor = db["products"].find({}, projection).limit(2000)

  async for doc in cursor:
    fields = doc.get("fields") or {}
    for field_name in PRODUCT_SELECTABLE_FIELDS:
      val = fields.get(field_name)
      values: list[str] = []
      if isinstance(val, str):
        values = [part.strip() for part in val.split(",")]
      elif isinstance(val, list):
        values = [part.strip() for part in val if isinstance(part, str)]

      for value in values:
        if value:
          options[field_name].add(value)

  return {
    field_name: sorted(values, key=lambda x: x.casefold())
    for field_name, values in options.items()
  }


async def _get_product_field_options(db: Any) -> dict[str, list[str]]:
  doc = await db[PRODUCT_FIELD_OPTIONS_COLLECTION].find_one({"_id": PRODUCT_FIELD_OPTIONS_DOC_ID})
  stored_options = doc.get("options") if isinstance(doc, dict) else None

  if isinstance(stored_options, dict):
    return _normalize_product_field_options_payload(stored_options)

  derived_options = await _derive_product_field_options(db)
  initial_options: dict[str, list[str]] = {}
  for field_name in PRODUCT_SELECTABLE_FIELDS:
    seen: set[str] = set()
    values: list[str] = []
    for value in [*DEFAULT_PRODUCT_FIELD_OPTIONS.get(field_name, []), *derived_options.get(field_name, [])]:
      key = value.casefold()
      if key in seen:
        continue
      seen.add(key)
      values.append(value)
    initial_options[field_name] = values

  return initial_options


def _normalize_import_header(value: Any) -> str:
  text = str(value or "").strip().casefold()
  text = re.sub(r"[_\-/:(){}\[\]#]+", " ", text)
  text = re.sub(r"\s+", " ", text)
  return text.strip()


def _canonical_import_column(header: Any, fallback: str) -> str:
  normalized = _normalize_import_header(header)
  if not normalized:
    return fallback
  return PRODUCT_IMPORT_HEADER_ALIASES.get(normalized) or str(header).strip()


def _serialize_excel_value(value: Any) -> Any:
  if value is None:
    return ""
  if isinstance(value, datetime):
    return value.isoformat()
  if isinstance(value, str):
    return value.strip()
  if isinstance(value, (int, float, bool)):
    return value
  return str(value).strip()


def _is_empty_excel_value(value: Any) -> bool:
  if value is None:
    return True
  if isinstance(value, str):
    return value.strip() == ""
  return False


def _score_import_header_row(values: tuple[Any, ...]) -> int:
  non_empty = 0
  alias_hits = 0
  for value in values:
    if _is_empty_excel_value(value):
      continue
    non_empty += 1
    if _normalize_import_header(value) in PRODUCT_IMPORT_HEADER_ALIASES:
      alias_hits += 1
  return alias_hits * 10 + min(non_empty, 8)


def _detect_import_header_row(rows: list[tuple[Any, ...]]) -> int:
  best_index = 0
  best_score = -1
  for idx, row in enumerate(rows[:30]):
    score = _score_import_header_row(row)
    if score > best_score:
      best_score = score
      best_index = idx
  return best_index


def _parse_import_price(value: Any) -> tuple[Any, str | None]:
  if value is None or value == "":
    return "", None
  if isinstance(value, (int, float)) and not isinstance(value, bool):
    return int(value) if float(value).is_integer() else value, None

  raw = str(value).strip()
  cleaned = re.sub(r"(?i)\b(aed|dirham|درهم)\b", "", raw)
  cleaned = cleaned.replace(",", "").strip()
  if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
    number = float(cleaned)
    return int(number) if number.is_integer() else number, None
  return raw, "Price could not be parsed as a clean number"


def _parse_import_bool(value: Any) -> Any:
  if isinstance(value, bool):
    return value
  if isinstance(value, (int, float)) and value in (0, 1):
    return bool(value)
  if isinstance(value, str):
    normalized = value.strip().casefold()
    if normalized in {"yes", "true", "1", "main", "y"}:
      return True
    if normalized in {"no", "false", "0", "n"}:
      return False
  return value


def _clean_import_field(column: str, value: Any) -> tuple[Any, str | None]:
  serialized = _serialize_excel_value(value)
  if column == "Price":
    return _parse_import_price(serialized)
  if column == "Main":
    return _parse_import_bool(serialized), None
  return serialized, None


def _merge_import_value(previous: Any, next_value: Any) -> Any:
  if previous in (None, ""):
    return next_value
  if next_value in (None, ""):
    return previous
  if previous == next_value:
    return previous
  return f"{previous}\n{next_value}"


def _import_value_to_text(value: Any) -> str:
  if value is None:
    return ""
  if isinstance(value, bool):
    return "true" if value else "false"
  if isinstance(value, (int, float)):
    return str(int(value)) if float(value).is_integer() else str(value)
  return str(value).strip()


def _trim_variant_fragment(value: str) -> str:
  return re.sub(r"^[\s\-_/\\|:.#]+|[\s\-_/\\|:.#]+$", "", value).strip()


def _normalized_fragment_positions(value: str) -> tuple[str, list[int]]:
  chars: list[str] = []
  positions: list[int] = []
  for idx, char in enumerate(value):
    if char.isalnum():
      chars.append(char.casefold())
      positions.append(idx)
  return "".join(chars), positions


def _find_subsequence_positions(value: str, fragment: str) -> list[int] | None:
  positions: list[int] = []
  search_from = 0
  for char in fragment:
    found_at = value.find(char, search_from)
    if found_at < 0:
      return None
    positions.append(found_at)
    search_from = found_at + 1
  return positions


def _derive_variant_number_from_code_number(code_number: Any, collection_code: Any) -> str | None:
  full_code = _import_value_to_text(code_number)
  base_code = _import_value_to_text(collection_code)
  if not full_code or not base_code:
    return None

  full_folded = full_code.casefold()
  base_folded = base_code.casefold()

  if full_folded.startswith(base_folded):
    candidate = _trim_variant_fragment(full_code[len(base_code):])
    return candidate or None

  base_match = re.search(re.escape(base_code), full_code, flags=re.IGNORECASE)
  if base_match:
    candidate = _trim_variant_fragment(full_code[:base_match.start()] + full_code[base_match.end():])
    return candidate or None

  normalized_full, full_positions = _normalized_fragment_positions(full_code)
  normalized_base, base_positions = _normalized_fragment_positions(base_code)
  if not normalized_full or not normalized_base or not base_positions:
    return None

  match_index = normalized_full.find(normalized_base)
  if match_index < 0:
    subsequence_positions = _find_subsequence_positions(normalized_full, normalized_base)
    if not subsequence_positions:
      return None

    original_positions_to_remove = {full_positions[idx] for idx in subsequence_positions}
    candidate = "".join(
      char for idx, char in enumerate(full_code)
      if idx not in original_positions_to_remove
    )
    candidate = _trim_variant_fragment(candidate)
    return candidate or None

  start = full_positions[match_index]
  end = full_positions[match_index + len(normalized_base) - 1] + 1
  candidate = _trim_variant_fragment(full_code[:start] + full_code[end:])
  return candidate or None


def _fill_collection_code_from_fallbacks(fields: dict[str, Any]) -> None:
  collection_code = _import_value_to_text(fields.get("Colecction Code"))
  if collection_code:
    return

  code_number = _import_value_to_text(fields.get("CODE NUMBER"))
  if code_number:
    fields["Colecction Code"] = code_number
    return

  factory_code = _import_value_to_text(fields.get("Factory Code"))
  if factory_code:
    fields["Colecction Code"] = factory_code


def _split_variant_from_collection_name(fields: dict[str, Any]) -> None:
  collection_name = _import_value_to_text(fields.get("Colecction Name"))
  if "/" not in collection_name:
    return

  name_part, variant_part = collection_name.rsplit("/", 1)
  clean_name = name_part.strip()
  clean_variant = _trim_variant_fragment(variant_part)
  if clean_name:
    fields["Colecction Name"] = clean_name
  if clean_variant:
    fields["Variant Number"] = clean_variant


def _cleanup_redundant_text(value: str) -> str:
  value = value.replace("\u00a0", " ")
  value = re.sub(r"[ \t]{2,}", " ", value)
  value = re.sub(r"\n{3,}", "\n\n", value)
  value = re.sub(r"\s*([,;|/])\s*([,;|/])+", r"\1", value)
  value = re.sub(r"^[\s,;|/\\._:-]+|[\s,;|/\\._:-]+$", "", value)
  return value.strip()


def _flexible_fragment_pattern(fragment: str) -> str:
  parts: list[str] = []
  for char in fragment.strip():
    if char.isspace() or char == "\u00a0":
      parts.append(r"\s+")
    elif char in {"*", "×", "x", "X"}:
      parts.append(r"\s*[*×xX]\s*")
    elif char in {"-", "_", "/", "\\", "|", ":", "."}:
      parts.append(r"\s*[-_/\\|:.]\s*")
    else:
      parts.append(re.escape(char))
  return "".join(parts)


def _remove_redundant_fragment(text: str, fragment: str) -> str:
  fragment = fragment.strip()
  if len(fragment) < 3:
    return text
  next_text = re.sub(_flexible_fragment_pattern(fragment), "", text, flags=re.IGNORECASE)
  return _cleanup_redundant_text(next_text)


def _remove_duplicate_details_content(fields: dict[str, Any]) -> None:
  details = _import_value_to_text(fields.get("Details"))
  if not details:
    return

  redundant_values = [
    _import_value_to_text(fields.get("DIMENSION (mm)")),
    _import_value_to_text(fields.get("Note")),
  ]

  next_details = details
  for redundant in redundant_values:
    if not redundant:
      continue
    next_details = _remove_redundant_fragment(next_details, redundant)
    for part in re.split(r"[\n;|]+", redundant):
      next_details = _remove_redundant_fragment(next_details, part)

  if next_details:
    fields["Details"] = next_details
  else:
    fields.pop("Details", None)


def _apply_import_row_formulas(fields: dict[str, Any]) -> dict[str, Any]:
  next_fields = dict(fields)
  preserved_images = {
    key: value
    for key, value in next_fields.items()
    if re.fullmatch(r"Image\d+", str(key))
  }

  _fill_collection_code_from_fallbacks(next_fields)

  derived_variant = _derive_variant_number_from_code_number(
    next_fields.get("CODE NUMBER"),
    next_fields.get("Colecction Code"),
  )
  if derived_variant:
    next_fields["Variant Number"] = derived_variant

  _split_variant_from_collection_name(next_fields)
  _remove_duplicate_details_content(next_fields)
  next_fields.update(preserved_images)
  return next_fields


def _normalize_import_row(headers: list[str], values: tuple[Any, ...]) -> tuple[dict[str, Any], dict[str, Any], list[str]]:
  fields: dict[str, Any] = {}
  raw_fields: dict[str, Any] = {}
  warnings: list[str] = []

  for idx, header in enumerate(headers):
    value = values[idx] if idx < len(values) else None
    serialized = _serialize_excel_value(value)
    if serialized == "":
      continue

    raw_fields[header] = _merge_import_value(raw_fields.get(header), serialized)
    cleaned_value, warning = _clean_import_field(header, serialized)
    fields[header] = _merge_import_value(fields.get(header), cleaned_value)
    if warning:
      warnings.append(f"{header}: {warning}")

  fields = _apply_import_row_formulas(fields)

  return fields, raw_fields, warnings


def _product_compare_value(value: Any) -> str:
  text = _import_value_to_text(value)
  if not text:
    return ""
  numeric = text.replace(",", "").strip()
  if re.fullmatch(r"-?\d+(\.\d+)?", numeric):
    number = float(numeric)
    return str(int(number)) if number.is_integer() else str(number)
  return re.sub(r"\s+", " ", text).casefold()


def _first_non_empty_field(fields: dict[str, Any], names: list[str]) -> Any:
  for name in names:
    value = fields.get(name)
    if _product_compare_value(value):
      return value
  return ""


def _product_match_keys(fields: dict[str, Any]) -> list[str]:
  collection_name = _product_compare_value(_first_non_empty_field(fields, ["Colecction Name", "Collection Name", "Name"]))
  collection_code = _product_compare_value(_first_non_empty_field(fields, ["Colecction Code", "Collection Code", "Code"]))
  variant = _product_compare_value(_first_non_empty_field(fields, ["Variant Number", "Variant", "Num"]))
  code_number = _product_compare_value(_first_non_empty_field(fields, ["CODE NUMBER", "Code Number", "Code No"]))

  return [
    f"code-number:{code_number}" if code_number else "",
    f"collection-code-variant:{collection_code}:{variant}" if collection_code and variant else "",
    f"collection-name-variant:{collection_name}:{variant}" if collection_name and variant else "",
    f"collection-code:{collection_code}" if collection_code else "",
  ]


def _canonical_product_field_value(fields: dict[str, Any], field_name: str) -> Any:
  aliases: dict[str, list[str]] = {
    "colecction name": ["Colecction Name", "Collection Name", "Name"],
    "collection name": ["Colecction Name", "Collection Name", "Name"],
    "name": ["Colecction Name", "Collection Name", "Name"],
    "colecction code": ["Colecction Code", "Collection Code", "Code"],
    "collection code": ["Colecction Code", "Collection Code", "Code"],
    "code": ["Colecction Code", "Collection Code", "Code"],
    "variant number": ["Variant Number", "Variant", "Num"],
    "variant": ["Variant Number", "Variant", "Num"],
    "num": ["Num", "Variant Number", "Variant"],
    "code number": ["CODE NUMBER", "Code Number", "Code No"],
    "dimension (mm)": ["DIMENSION (mm)", "Dimension (mm)", "DIMENSION (cm)", "Dimension (cm)", "DIMENSION", "Dimension", "Dimensions", "Size"],
    "dimension (cm)": ["DIMENSION (mm)", "Dimension (mm)", "DIMENSION (cm)", "Dimension (cm)", "DIMENSION", "Dimension", "Dimensions", "Size"],
  }
  names = aliases.get(field_name.strip().casefold(), [field_name])
  return _first_non_empty_field(fields, names)


async def _build_product_match_index(db: Any) -> dict[str, dict[str, Any]]:
  index: dict[str, dict[str, Any]] = {}
  cursor = db["products"].find({})
  async for doc in cursor:
    fields = doc.get("fields") or {}
    if not isinstance(fields, dict):
      continue
    for key in _product_match_keys(fields):
      if key and key not in index:
        index[key] = doc
  return index


def _find_matching_product(row_fields: dict[str, Any], product_index: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
  for key in _product_match_keys(row_fields):
    product = product_index.get(key)
    if product:
      return product
  return None


def _validate_import_match_column(column: str) -> str:
  name = column.strip()
  if not name or "." in name or name.startswith("$"):
    raise HTTPException(status_code=400, detail="INVALID_MATCH_COLUMN")
  return name


def _import_row_field_value(fields: dict[str, Any], column: str) -> Any:
  if column in fields:
    return fields[column]
  return _canonical_product_field_value(fields, column)


async def _build_product_index_by_column(db: Any, product_column: str) -> dict[str, dict[str, Any]]:
  index: dict[str, dict[str, Any]] = {}
  cursor = db["products"].find({})
  async for doc in cursor:
    fields = doc.get("fields") or {}
    if not isinstance(fields, dict):
      continue
    raw_value = _canonical_product_field_value(fields, product_column)
    key = _product_compare_value(raw_value)
    if key and key not in index:
      index[key] = doc
  return index


def _find_matching_product_by_column(
  row_fields: dict[str, Any],
  product_index: dict[str, dict[str, Any]],
  import_column: str,
) -> dict[str, Any] | None:
  raw_value = _import_row_field_value(row_fields, import_column)
  key = _product_compare_value(raw_value)
  if not key:
    return None
  return product_index.get(key)


def _parse_import_match_payload(payload: dict[str, Any] | None) -> tuple[str, str] | None:
  if not isinstance(payload, dict):
    return None
  match = payload.get("match")
  if not isinstance(match, dict):
    return None
  import_column = match.get("import_column")
  product_column = match.get("product_column")
  if not isinstance(import_column, str) or not isinstance(product_column, str):
    return None
  return _validate_import_match_column(import_column), _validate_import_match_column(product_column)


PRODUCT_IMPORT_ROW_MATCH_STATUSES = frozenset({"matched", "unmatched", "empty"})


def _classify_import_row_match_status(
  row_fields: dict[str, Any],
  product_index: dict[str, dict[str, Any]],
  import_column: str,
) -> str:
  if not isinstance(row_fields, dict):
    return "empty"
  import_value = _import_row_field_value(row_fields, import_column)
  import_key = _product_compare_value(import_value)
  if not import_key:
    return "empty"
  if product_index.get(import_key):
    return "matched"
  return "unmatched"


def _parse_apply_row_groups(payload: dict[str, Any] | None) -> set[str]:
  if not isinstance(payload, dict):
    return set(PRODUCT_IMPORT_ROW_MATCH_STATUSES)
  raw = payload.get("apply_row_groups")
  if raw is None:
    return set(PRODUCT_IMPORT_ROW_MATCH_STATUSES)
  if not isinstance(raw, list) or not raw:
    raise HTTPException(status_code=400, detail="INVALID_APPLY_ROW_GROUPS")
  groups: set[str] = set()
  for item in raw:
    if isinstance(item, str):
      normalized = item.strip()
      if normalized in PRODUCT_IMPORT_ROW_MATCH_STATUSES:
        groups.add(normalized)
  if not groups:
    raise HTTPException(status_code=400, detail="INVALID_APPLY_ROW_GROUPS")
  return groups


def _product_image_url_to_storage_path(url: Any) -> Path | None:
  value = _import_value_to_text(url)
  if not value:
    return None

  prefixes = ["/api/trainer/files/", "/files/"]
  relative = ""
  for prefix in prefixes:
    if value.startswith(prefix):
      relative = value[len(prefix):]
      break
  if not relative:
    return None

  path = (STORAGE_DIR / relative).resolve()
  storage_root = STORAGE_DIR.resolve()
  try:
    path.relative_to(storage_root)
  except ValueError:
    return None
  return path


def _copy_import_image_to_product_storage(image_url: Any, product_id: str) -> str:
  source = _product_image_url_to_storage_path(image_url)
  if not source or not source.exists():
    return _import_value_to_text(image_url)

  target_dir = STORAGE_DIR / "product-images" / product_id
  target_dir.mkdir(parents=True, exist_ok=True)
  suffix = source.suffix or ".png"
  target_name = f"{uuid.uuid4().hex}{suffix}"
  target = target_dir / target_name
  shutil.copyfile(source, target)
  return f"/api/trainer/files/product-images/{product_id}/{target_name}"


def _product_fields_from_import_fields(
  fields: dict[str, Any],
  product_id: str,
  existing_fields: dict[str, Any] | None = None,
  selected_columns: set[str] | None = None,
) -> dict[str, Any]:
  product_fields: dict[str, Any] = {}
  existing_image = _product_compare_value((existing_fields or {}).get("Image"))

  for field_name, value in fields.items():
    if field_name == "Row":
      if selected_columns is not None and "Row" not in selected_columns:
        continue
      if value != "":
        product_fields["Num"] = value
      continue
    if field_name == "Image1":
      continue
    if selected_columns is not None and field_name not in selected_columns:
      continue
    if value == "":
      continue
    product_fields[field_name] = value

  image1 = fields.get("Image1")
  image1_selected = selected_columns is None or "Image1" in selected_columns
  if image1_selected and _product_compare_value(image1) and not existing_image:
    product_fields["Image"] = _copy_import_image_to_product_storage(image1, product_id)

  return product_fields


def _changed_product_fields(
  staged_fields: dict[str, Any],
  existing_fields: dict[str, Any],
  product_id: str,
  selected_columns: set[str] | None = None,
) -> dict[str, Any]:
  candidate_fields = _product_fields_from_import_fields(staged_fields, product_id, existing_fields, selected_columns)
  changed: dict[str, Any] = {}

  for field_name, value in candidate_fields.items():
    if "." in field_name or field_name.startswith("$"):
      continue
    existing_value = _canonical_product_field_value(existing_fields, field_name)
    if _product_compare_value(existing_value) == _product_compare_value(value):
      continue
    changed[field_name] = value

  return changed


def _safe_import_filename(filename: str, default: str) -> str:
  name = Path(filename or default).name
  name = re.sub(r"[^A-Za-z0-9._ -]+", "_", name).strip(" .")
  return name or default


def _run_numbers_conversion_with_libreoffice(input_path: Path, output_dir: Path) -> bytes | None:
  converter = shutil.which("soffice") or shutil.which("libreoffice")
  if not converter:
    return None

  result = subprocess.run(
    [
      converter,
      "--headless",
      "--convert-to",
      "xlsx",
      "--outdir",
      str(output_dir),
      str(input_path),
    ],
    check=False,
    capture_output=True,
    text=True,
    timeout=120,
  )
  if result.returncode != 0:
    raise HTTPException(status_code=400, detail=f"NUMBERS_CONVERSION_FAILED: {result.stderr or result.stdout}")

  candidates = sorted(output_dir.glob("*.xlsx"))
  if not candidates:
    raise HTTPException(status_code=400, detail="NUMBERS_CONVERSION_FAILED: converted xlsx not found")
  return candidates[0].read_bytes()


def _run_numbers_conversion_with_macos_numbers(input_path: Path, output_path: Path) -> bytes | None:
  osascript = shutil.which("osascript")
  if not osascript:
    return None

  script = """
on run argv
  set inputPath to POSIX file (item 1 of argv)
  set outputPath to POSIX file (item 2 of argv)
  tell application "Numbers"
    set theDoc to open inputPath
    export theDoc to outputPath as Microsoft Excel
    close theDoc saving no
  end tell
end run
"""
  result = subprocess.run(
    [osascript, "-e", script, str(input_path), str(output_path)],
    check=False,
    capture_output=True,
    text=True,
    timeout=120,
  )
  if result.returncode != 0:
    raise HTTPException(status_code=400, detail=f"NUMBERS_CONVERSION_FAILED: {result.stderr or result.stdout}")
  if not output_path.exists():
    raise HTTPException(status_code=400, detail="NUMBERS_CONVERSION_FAILED: converted xlsx not found")
  return output_path.read_bytes()


def _convert_numbers_to_xlsx_content(content: bytes, filename: str) -> bytes:
  with tempfile.TemporaryDirectory(prefix="product-import-") as tmp_dir:
    tmp_path = Path(tmp_dir)
    input_path = tmp_path / _safe_import_filename(filename, "products-import.numbers")
    if input_path.suffix.lower() != ".numbers":
      input_path = input_path.with_suffix(".numbers")
    input_path.write_bytes(content)

    output_dir = tmp_path / "converted"
    output_dir.mkdir(parents=True, exist_ok=True)

    converted = _run_numbers_conversion_with_libreoffice(input_path, output_dir)
    if converted is not None:
      return converted

    converted = _run_numbers_conversion_with_macos_numbers(input_path, output_dir / "converted.xlsx")
    if converted is not None:
      return converted

  raise HTTPException(
    status_code=400,
    detail="NUMBERS_CONVERSION_UNAVAILABLE: install LibreOffice or run the server on macOS with Numbers available",
  )


def _image_extension(image: Any) -> str:
  image_format = str(getattr(image, "format", "") or "").lower().strip(".")
  if image_format in {"jpeg", "jpg"}:
    return "jpg"
  if image_format in {"png", "gif", "webp", "bmp"}:
    return image_format

  path = str(getattr(image, "path", "") or "")
  suffix = Path(path).suffix.lower().strip(".")
  if suffix in {"jpeg", "jpg", "png", "gif", "webp", "bmp"}:
    return "jpg" if suffix == "jpeg" else suffix
  return "png"


def _extract_image_bytes(image: Any) -> bytes | None:
  try:
    data = image._data()
  except Exception:
    return None
  return data if isinstance(data, bytes) and data else None


def _worksheet_image_anchor(image: Any) -> tuple[int, int] | None:
  marker = getattr(getattr(image, "anchor", None), "_from", None)
  if marker is None:
    return None
  row = getattr(marker, "row", None)
  col = getattr(marker, "col", None)
  if not isinstance(row, int) or not isinstance(col, int):
    return None
  return row + 1, col + 1


def _extract_first_column_images(worksheet: Any, import_id: str) -> dict[int, list[str]]:
  images_by_row: dict[int, list[str]] = {}
  images = list(getattr(worksheet, "_images", []) or [])
  if not images:
    return images_by_row

  target_dir = STORAGE_DIR / "product-imports" / import_id
  target_dir.mkdir(parents=True, exist_ok=True)
  safe_sheet = re.sub(r"[^A-Za-z0-9._-]+", "_", str(worksheet.title or "sheet")).strip("_") or "sheet"

  for index, image in enumerate(images, start=1):
    anchor = _worksheet_image_anchor(image)
    if not anchor:
      continue
    row_number, column_number = anchor
    if column_number != 1:
      continue

    data = _extract_image_bytes(image)
    if not data:
      continue

    ext = _image_extension(image)
    filename = f"{safe_sheet}_r{row_number}_{index}.{ext}"
    (target_dir / filename).write_bytes(data)
    images_by_row.setdefault(row_number, []).append(f"/api/trainer/files/product-imports/{import_id}/{filename}")

  return images_by_row


def _parse_product_import_workbook(content: bytes, filename: str, import_id: str) -> dict[str, Any]:
  try:
    from openpyxl import load_workbook
  except ImportError as exc:
    raise HTTPException(status_code=500, detail="OPENPYXL_NOT_INSTALLED") from exc

  workbook_content = _convert_numbers_to_xlsx_content(content, filename) if filename.lower().endswith(".numbers") else content

  try:
    workbook = load_workbook(io.BytesIO(workbook_content), read_only=False, data_only=True)
  except Exception as exc:
    raise HTTPException(status_code=400, detail=f"INVALID_EXCEL_FILE: {exc}") from exc

  rows_to_insert: list[dict[str, Any]] = []
  columns_seen: set[str] = set(PRODUCT_IMPORT_CANONICAL_COLUMNS)
  sheets: list[dict[str, Any]] = []

  for worksheet in workbook.worksheets:
    sheet_rows = list(worksheet.iter_rows(values_only=True))
    if not sheet_rows:
      continue
    images_by_row = _extract_first_column_images(worksheet, import_id)

    header_index = _detect_import_header_row(sheet_rows)
    header_values = sheet_rows[header_index]
    headers: list[str] = []
    used_headers: dict[str, int] = {}

    for idx, value in enumerate(header_values):
      fallback = f"Column {idx + 1}"
      header = _canonical_import_column(value, fallback)
      if header in used_headers:
        used_headers[header] += 1
        header = f"{header} {used_headers[header]}"
      else:
        used_headers[header] = 1
      headers.append(header)

    sheet_count = 0
    for row_offset, values in enumerate(sheet_rows[header_index + 1:], start=header_index + 2):
      if all(_is_empty_excel_value(value) for value in values) and not images_by_row.get(row_offset):
        continue

      fields, raw_fields, warnings = _normalize_import_row(headers, values)
      for image_index, image_url in enumerate(images_by_row.get(row_offset, []), start=1):
        image_field = f"Image{image_index}"
        fields[image_field] = image_url
        raw_fields[image_field] = image_url

      if not fields:
        continue

      for key in fields.keys():
        columns_seen.add(key)

      rows_to_insert.append({
        "_id": uuid.uuid4().hex,
        "source_sheet": worksheet.title,
        "source_row_number": row_offset,
        "fields": fields,
        "raw_fields": raw_fields,
        "warnings": warnings,
        "status": "staged",
      })
      sheet_count += 1

    sheets.append({
      "name": worksheet.title,
      "header_row": header_index + 1,
      "row_count": sheet_count,
    })

  ordered_columns = [
    *[column for column in PRODUCT_IMPORT_CANONICAL_COLUMNS if column in columns_seen],
    *sorted((column for column in columns_seen if column not in PRODUCT_IMPORT_CANONICAL_COLUMNS), key=lambda x: x.casefold()),
  ]

  return {
    "columns": ordered_columns,
    "rows": rows_to_insert,
    "sheets": sheets,
  }

class ClassItem(TypedDict):
  id: str
  name: str


class NormalizedBBox(TypedDict):
  x: float
  y: float
  w: float
  h: float


class Annotation(TypedDict):
  class_id: str
  bbox: NormalizedBBox


QueueStatus = Literal["pending", "labeled"]


class QueueItem(TypedDict, total=False):
  item_id: str
  filename: str
  status: QueueStatus
  created_at: str
  annotation: Annotation


pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")


def _env_int(name: str, default: int) -> int:
  raw = os.environ.get(name)
  if not raw:
    return default
  try:
    return int(raw)
  except Exception:
    return default


def _env_str(name: str, default: str) -> str:
  raw = os.environ.get(name)
  return raw if raw else default


def _local_dev_cors_origins() -> list[str]:
  return [
    "http://localhost:3010",
    "http://127.0.0.1:3010",
    "http://localhost:3004",
    "http://127.0.0.1:3004",
    "http://localhost:3003",
    "http://127.0.0.1:3003",
    "http://localhost:3005",
    "http://127.0.0.1:3005",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
  ]


def _cors_allow_origins() -> list[str]:
  raw = os.environ.get("TRAINER_CORS_ORIGINS")
  if raw:
    origins = [x.strip() for x in raw.split(",") if x.strip()]
    return origins + _local_dev_cors_origins()

  domain = os.environ.get("APP_BASE_DOMAIN", "ehsanrahimi.com").strip() or "ehsanrahimi.com"
  hub_sub = os.environ.get("HUB_SUBDOMAIN", "dashboard").strip() or "dashboard"
  production = [
    f"https://{hub_sub}.{domain}",
    f"https://products.{domain}",
    f"https://marketing.{domain}",
    f"https://trainer.{domain}",
  ]
  return production + _local_dev_cors_origins()


def _is_production() -> bool:
  return (os.environ.get("ENV") == "production") or (os.environ.get("NODE_ENV") == "production")


def _hash_password(password: str) -> str:
  return pwd_context.hash(password)


def _verify_password(password: str, password_hash: str) -> bool:
  try:
    return pwd_context.verify(password, password_hash)
  except Exception:
    return False


def _jwt_secret() -> str:
  secret = os.environ.get("TRAINER_JWT_SECRET")
  if not secret:
    raise HTTPException(status_code=500, detail="TRAINER_JWT_SECRET_NOT_SET")
  if len(secret.strip()) < 16:
    raise HTTPException(status_code=500, detail="TRAINER_JWT_SECRET_TOO_SHORT")
  return secret


def _auth_cookie_name() -> str:
  return _env_str("TRAINER_AUTH_COOKIE_NAME", "trainer_auth")


def _cookie_domain() -> str | None:
  raw = os.environ.get("TRAINER_COOKIE_DOMAIN")
  if not raw:
    return None
  v = raw.strip()
  return v if v else None


def _admin_email_norm() -> str:
  raw = os.environ.get("TRAINER_ADMIN_EMAIL")
  v = raw.strip().lower() if isinstance(raw, str) else ""
  return v if v else "ehsanrahimi8@gmail.com"


def _normalize_role(value: Any) -> str:
  if not isinstance(value, str):
    return "user"
  v = value.strip().lower()
  if v in ["user", "sales", "admin"]:
    return v
  return "user"


def _normalize_permissions(value: Any) -> list[str]:
  if not isinstance(value, list):
    return []
  out: list[str] = []
  for item in value:
    if isinstance(item, str) and item.strip():
      out.append(item.strip())
  return sorted(set(out))


def _create_access_token(*, user_id: str, is_admin: bool, permissions: list[str]) -> str:
  now = int(time.time())
  exp_s = _env_int("TRAINER_JWT_EXPIRES_SECONDS", 60 * 60 * 24 * 7)
  payload = {
    "sub": user_id,
    "iat": now,
    "exp": now + exp_s,
    "is_admin": bool(is_admin),
    "permissions": permissions,
  }
  return jwt.encode(payload, _jwt_secret(), algorithm="HS256")


def _decode_access_token(token: str) -> dict[str, Any]:
  try:
    decoded = jwt.decode(token, _jwt_secret(), algorithms=["HS256"])
  except jwt.ExpiredSignatureError:
    raise HTTPException(status_code=401, detail="TOKEN_EXPIRED")
  except Exception:
    raise HTTPException(status_code=401, detail="INVALID_TOKEN")
  if not isinstance(decoded, dict):
    raise HTTPException(status_code=401, detail="INVALID_TOKEN")
  return decoded


def _utc_now_iso() -> str:
  return datetime.now(timezone.utc).isoformat()


_CONTENT_CALENDAR_FIELDS = [
  "Title",
  "Publish Date",
  "Day of Week",
  "Content Pillar",
  "Format",
  "Status",
  "Content Link",
  "Social Views",
  "Caption Idea",
  "CTA",
  "Tone of Voice",
  "Target Audience",
  "# Hashtag",
  "Product",
  "Product Image",
  "Assets",
  "_campaign_planning_id",
]


def _sanitize_content_calendar_fields(raw: Any) -> dict[str, Any]:
  if not isinstance(raw, dict):
    return {}
  out: dict[str, Any] = {}
  for k in _CONTENT_CALENDAR_FIELDS:
    if k in raw:
      out[k] = raw.get(k)
  return out


def _parse_multi_value_field(raw: Any) -> list[str]:
  if not isinstance(raw, str):
    return []
  return [part.strip() for part in re.split(r"[,;\n]+", raw) if part.strip()]


def _serialize_multi_value_field(values: list[str]) -> str:
  return ", ".join(values)


def _remove_multi_value_item(raw: Any, item: str) -> str:
  target = item.strip().casefold()
  if not target:
    return _serialize_multi_value_field(_parse_multi_value_field(raw))
  next_values = [v for v in _parse_multi_value_field(raw) if v.strip().casefold() != target]
  return _serialize_multi_value_field(next_values)


def _normalize_string_option_list(raw: Any) -> list[str]:
  if not isinstance(raw, list):
    return []
  seen: set[str] = set()
  cleaned: list[str] = []
  for item in raw:
    if not isinstance(item, str):
      continue
    value = item.strip()
    key = value.casefold()
    if not value or key in seen:
      continue
    seen.add(key)
    cleaned.append(value)
  return sorted(cleaned, key=lambda x: x.casefold())


async def _derive_calendar_field_options(db: Any, field_name: str) -> list[str]:
  seen: set[str] = set()
  values: list[str] = []
  projection = {f"fields.{field_name}": 1}
  cursor = db["content_calendar"].find({}, projection)

  async for doc in cursor:
    fields = doc.get("fields") if isinstance(doc.get("fields"), dict) else {}
    raw = fields.get(field_name)
    parts: list[str] = []
    if field_name in CONTENT_CALENDAR_MULTI_VALUE_FIELDS:
      parts = _parse_multi_value_field(raw)
    elif isinstance(raw, str) and raw.strip():
      parts = [raw.strip()]

    for part in parts:
      key = part.casefold()
      if key in seen:
        continue
      seen.add(key)
      values.append(part)

  return sorted(values, key=lambda x: x.casefold())


async def _get_stored_calendar_field_options(db: Any) -> dict[str, list[str]]:
  doc = await db[CONTENT_CALENDAR_FIELD_OPTIONS_COLLECTION].find_one({"_id": CONTENT_CALENDAR_FIELD_OPTIONS_DOC_ID})
  stored_options = doc.get("options") if isinstance(doc, dict) else None
  if not isinstance(stored_options, dict):
    return {}

  out: dict[str, list[str]] = {}
  for field_name in CONTENT_CALENDAR_SELECTABLE_FIELDS:
    raw_values = stored_options.get(field_name)
    if isinstance(raw_values, list):
      out[field_name] = _normalize_string_option_list(raw_values)
  return out


async def _set_calendar_field_options_for_field(db: Any, field_name: str, values: list[str]) -> list[str]:
  if field_name not in CONTENT_CALENDAR_SELECTABLE_FIELDS:
    raise HTTPException(status_code=400, detail="INVALID_FIELD")

  normalized = _normalize_string_option_list(values)
  now = _utc_now_iso()
  stored = await _get_stored_calendar_field_options(db)
  stored[field_name] = normalized

  await db[CONTENT_CALENDAR_FIELD_OPTIONS_COLLECTION].update_one(
    {"_id": CONTENT_CALENDAR_FIELD_OPTIONS_DOC_ID},
    {
      "$set": {
        "options": stored,
        "updated_at": now,
      }
    },
    upsert=True,
  )
  return normalized


async def _get_calendar_field_options(db: Any, field_name: str) -> list[str]:
  if field_name not in CONTENT_CALENDAR_SELECTABLE_FIELDS:
    raise HTTPException(status_code=400, detail="INVALID_FIELD")

  stored = await _get_stored_calendar_field_options(db)
  if field_name in stored and stored[field_name]:
    return stored[field_name]

  derived = await _derive_calendar_field_options(db, field_name)
  defaults = DEFAULT_CALENDAR_FIELD_OPTIONS.get(field_name, [])
  seen: set[str] = set()
  merged: list[str] = []
  for value in [*defaults, *derived]:
    key = value.casefold()
    if key in seen:
      continue
    seen.add(key)
    merged.append(value)
  return await _set_calendar_field_options_for_field(db, field_name, merged)


async def _get_all_calendar_field_options(db: Any) -> dict[str, list[str]]:
  out: dict[str, list[str]] = {}
  for field_name in CONTENT_CALENDAR_SELECTABLE_FIELDS:
    out[field_name] = await _get_calendar_field_options(db, field_name)
  return out


async def _add_calendar_field_option(db: Any, field_name: str, raw_value: str) -> list[str]:
  value = raw_value.strip()
  if not value:
    raise HTTPException(status_code=400, detail="VALUE_REQUIRED")

  current = await _get_calendar_field_options(db, field_name)
  if any(v.casefold() == value.casefold() for v in current):
    return current

  return await _set_calendar_field_options_for_field(db, field_name, [*current, value])


async def _remove_calendar_field_option_from_registry(db: Any, field_name: str, raw_value: str) -> list[str]:
  value = raw_value.strip()
  if not value:
    raise HTTPException(status_code=400, detail="VALUE_REQUIRED")

  current = await _get_calendar_field_options(db, field_name)
  next_values = [v for v in current if v.casefold() != value.casefold()]
  return await _set_calendar_field_options_for_field(db, field_name, next_values)


def _normalize_calendar_status_value(raw: Any) -> str:
  if not isinstance(raw, str):
    return ""
  status = raw.strip()
  if not status:
    return ""
  return "Drafts" if status == "Draft" else status


def _calendar_field_value_matches(raw: Any, value: str, field_name: str) -> bool:
  target = value.strip().casefold()
  if not target:
    return False

  if field_name in CONTENT_CALENDAR_MULTI_VALUE_FIELDS:
    return any(part.strip().casefold() == target for part in _parse_multi_value_field(raw))

  if field_name == "Status":
    return _normalize_calendar_status_value(raw).casefold() == target

  return isinstance(raw, str) and raw.strip().casefold() == target


def _remove_calendar_field_value(raw: Any, value: str, field_name: str) -> str:
  if field_name in CONTENT_CALENDAR_MULTI_VALUE_FIELDS:
    return _remove_multi_value_item(raw, value)

  if _calendar_field_value_matches(raw, value, field_name):
    return ""
  return raw if isinstance(raw, str) else ""


async def _remove_calendar_field_option_from_all_cells(
  db: Any,
  field_name: str,
  value: str,
  user_id: Any,
) -> int:
  now = _utc_now_iso()
  updated_items = 0

  cursor = db["content_calendar"].find({}, {"_id": 1, "fields": 1})
  async for doc in cursor:
    fields = doc.get("fields") if isinstance(doc.get("fields"), dict) else {}
    current = fields.get(field_name)
    if current is None:
      continue

    next_value = _remove_calendar_field_value(current, value, field_name)
    if next_value == current:
      continue

    next_fields = {**fields, field_name: next_value}
    await db["content_calendar"].update_one(
      {"_id": doc["_id"]},
      {
        "$set": {
          "fields": next_fields,
          "updated_at": now,
          "updated_by": user_id,
        }
      },
    )
    updated_items += 1

  return updated_items


def _safe_write_json(path: Path, data: Any) -> None:
  tmp = path.with_suffix(path.suffix + ".tmp")
  tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
  tmp.replace(path)


def _parse_user_agent(ua: str | None) -> str:
  if not ua:
    return "Unknown"
  ua_lower = ua.lower()
  os_name = "Unknown OS"
  if "windows" in ua_lower:
    os_name = "Windows"
  elif "android" in ua_lower:
    os_name = "Android"
  elif "iphone" in ua_lower or "ipad" in ua_lower:
    os_name = "iOS"
  elif "macintosh" in ua_lower or "mac os x" in ua_lower:
    os_name = "macOS"
  elif "linux" in ua_lower:
    os_name = "Linux"

  browser = "Unknown Browser"
  if "edg/" in ua_lower:
    browser = "Edge"
  elif "chrome/" in ua_lower:
    browser = "Chrome"
  elif "safari/" in ua_lower:
    browser = "Safari"
  elif "firefox/" in ua_lower:
    browser = "Firefox"
  elif "opr/" in ua_lower or "opera" in ua_lower:
    browser = "Opera"

  return f"{browser} on {os_name}"


def _serialize_field_value(value: Any) -> str:
  if value is None:
    return ""
  if isinstance(value, (dict, list)):
    try:
      return json.dumps(value, ensure_ascii=False)
    except Exception:
      return str(value)
  return str(value)


def _infer_field_change_type(old_value: str, new_value: str) -> str:
  if not old_value and new_value:
    return "add"
  if old_value and not new_value:
    return "clear"
  if old_value != new_value:
    return "update"
  return "unchanged"


def _build_field_change_entries_from_patch(
  existing_fields: dict[str, Any],
  changed_fields: dict[str, Any],
) -> list[dict[str, Any]]:
  entries: list[dict[str, Any]] = []
  for field_name, new_value in changed_fields.items():
    old_v = _canonical_product_field_value(existing_fields, field_name)
    old_s = _serialize_field_value(old_v)
    new_s = _serialize_field_value(new_value)
    if old_s == new_s:
      continue
    entries.append(
      {
        "field": field_name,
        "old_value": old_s,
        "new_value": new_s,
        "change_type": _infer_field_change_type(old_s, new_s),
      }
    )
  return entries


def _build_field_change_entries_from_create(new_fields: dict[str, Any]) -> list[dict[str, Any]]:
  entries: list[dict[str, Any]] = []
  for field_name, value in new_fields.items():
    new_s = _serialize_field_value(value)
    if not new_s:
      continue
    entries.append(
      {
        "field": field_name,
        "old_value": "",
        "new_value": new_s,
        "change_type": "add",
      }
    )
  return entries


async def _insert_import_apply_field_audit_logs(
  req: FastAPIRequest,
  *,
  user: dict[str, Any],
  db: Any,
  import_label: str,
  pending: list[tuple[str, list[dict[str, Any]], str]],
) -> None:
  if not pending:
    return

  now = _utc_now_iso()
  ip = req.client.host if req.client else "Unknown"
  ua = req.headers.get("user-agent", "")
  device = _parse_user_agent(ua)
  user_id = user.get("_id") if user else "anonymous"
  username = user.get("username") if user else "anonymous"

  docs: list[dict[str, Any]] = []
  for product_id, field_changes, kind in pending:
    if not field_changes:
      continue
    docs.append(
      {
        "timestamp": now,
        "user_id": user_id,
        "username": username,
        "action": "PRODUCT_IMPORT_EDIT",
        "resource_id": product_id,
        "details": f"Excel import «{import_label}» ({kind})",
        "ip_address": ip,
        "user_agent": ua,
        "device": device,
        "field_changes": field_changes,
      }
    )

  chunk_size = 500
  for offset in range(0, len(docs), chunk_size):
    await db["activity_logs"].insert_many(docs[offset : offset + chunk_size])


_PRODUCT_EDIT_FIELDS_RE = re.compile(r"Fields:\s*([^.]+)", re.IGNORECASE)
_PRODUCT_INLINE_FIELDS_RE = re.compile(r"Updated fields:\s*([^.]+)", re.IGNORECASE)


def _parse_product_edit_fields_from_details(details: str) -> list[str]:
  if not details:
    return []
  match = _PRODUCT_EDIT_FIELDS_RE.search(details) or _PRODUCT_INLINE_FIELDS_RE.search(details)
  if not match:
    return []
  return [part.strip() for part in match.group(1).split(",") if part.strip()]


async def log_activity(
  req: FastAPIRequest,
  action: str,
  details: str = "",
  resource_id: str | None = None,
  user: dict[str, Any] | None = None,
  db: Any = None,
  field_changes: list[dict[str, Any]] | None = None,
):
  if db is None:
    return

  now = _utc_now_iso()
  ip = req.client.host if req.client else "Unknown"
  ua = req.headers.get("user-agent", "")
  device = _parse_user_agent(ua)

  doc = {
    "timestamp": now,
    "user_id": user.get("_id") if user else "anonymous",
    "username": user.get("username") if user else "anonymous",
    "action": action,
    "resource_id": resource_id,
    "details": details,
    "ip_address": ip,
    "user_agent": ua,
    "device": device,
  }
  if field_changes:
    doc["field_changes"] = field_changes

  try:
    print(f"[Logging] Attempting to log: {action} by {user.get('email') if user else 'anonymous'}", flush=True)
    await db["activity_logs"].insert_one(doc)
    print(f"[Logging] ✓ Success", flush=True)
  except Exception as e:
    print(f"[Logging] ✗ Error: {e}", flush=True)


def _read_json(path: Path, default: Any) -> Any:
  if not path.exists():
    return default
  try:
    return json.loads(path.read_text(encoding="utf-8"))
  except Exception:
    return default





def _validate_class_id(value: str) -> None:
  if not value:
    raise HTTPException(status_code=400, detail="INVALID_CLASS_ID")
  if not re.fullmatch(r"[a-z0-9_-]+", value):
    raise HTTPException(status_code=400, detail="INVALID_CLASS_ID")


api = FastAPI(title="Lorenzo Trainer Server")

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
STATIC_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(docs_url=None, redoc_url=None, openapi_url=None)
api.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/trainer/api", api)
app.mount("/api", api)
app.mount("/", api)


def _get_db():
  db = getattr(app.state, "mongo_db", None)
  if db is None:
    raise HTTPException(status_code=503, detail="MONGODB_NOT_CONFIGURED")
  return db


async def _get_current_user(req: FastAPIRequest, db=Depends(_get_db)) -> dict[str, Any]:
  token = None
  auth_header = req.headers.get("authorization")
  if auth_header and auth_header.lower().startswith("bearer "):
    token = auth_header.split(" ", 1)[1].strip()
  if not token:
    token = req.cookies.get(_auth_cookie_name())
  if not token:
    raise HTTPException(status_code=401, detail="NOT_AUTHENTICATED")

  decoded = _decode_access_token(token)
  user_id = decoded.get("sub")
  if not isinstance(user_id, str) or not user_id:
    raise HTTPException(status_code=401, detail="INVALID_TOKEN")

  user = await db["users"].find_one({"_id": user_id})
  if not user:
    raise HTTPException(status_code=401, detail="USER_NOT_FOUND")
  if user.get("status") != "approved":
    raise HTTPException(status_code=403, detail="USER_NOT_APPROVED")

  user["permissions"] = _normalize_permissions(user.get("permissions"))
  print(f"[Auth] Request from: {user.get('email')} (ID: {user_id})", flush=True)
  return user


async def _require_admin(user: dict[str, Any] = Depends(_get_current_user)) -> dict[str, Any]:
  if user.get("is_admin") is True:
    return user
  raise HTTPException(status_code=403, detail="ADMIN_ONLY")


def _require_role(user: dict[str, Any], allowed_roles: set[str]) -> dict[str, Any]:
  role = _normalize_role(user.get("role"))
  if role in allowed_roles:
    return user
  raise HTTPException(status_code=403, detail="FORBIDDEN_ROLE")


async def _require_operator(user: dict[str, Any] = Depends(_get_current_user)) -> dict[str, Any]:
  return _require_role(user, {"admin", "sales"})

api.add_middleware(
  CORSMiddleware,
  allow_origins=_cors_allow_origins(),
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
)

try:
  from dotenv import load_dotenv  # type: ignore

  load_dotenv(dotenv_path=BASE_DIR / ".env", override=False)
except Exception:
  pass


_log = logging.getLogger("uvicorn.error")


@app.on_event("startup")
async def _startup_mongo_after_env_loaded():
  uri = os.environ.get("MONGODB_URI")
  db_name = _env_str("MONGODB_DB_NAME", "trainer")

  if not uri:
    app.state.mongo_client = None
    app.state.mongo_db = None
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
    print("⚠  [MongoDB] MONGODB_URI is not set in .env", flush=True)
    print("   Auth endpoints (login/register/me) will return 503.", flush=True)
    print("   Fix: set MONGODB_URI, TRAINER_JWT_SECRET, TRAINER_ADMIN_EMAIL", flush=True)
    print("        in trainer/server/.env", flush=True)
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
    return

  uri_display = uri[:40] + "..." if len(uri) > 40 else uri
  print(f"[MongoDB] Connecting to db='{db_name}' uri={uri_display}", flush=True)

  try:
    client = AsyncIOMotorClient(uri, serverSelectionTimeoutMS=8000)
    await client.admin.command("ping")
  except Exception as e:
    app.state.mongo_client = None
    app.state.mongo_db = None
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
    print(f"✗  [MongoDB] Connection FAILED: {e}", flush=True)
    print("   Check: MONGODB_URI value, Atlas Network Access (IP whitelist),", flush=True)
    print("          and database user credentials.", flush=True)
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", flush=True)
    return

  app.state.mongo_client = client
  app.state.mongo_db = client[db_name]
  print("[MongoDB] ✓ Connected successfully.", flush=True)

  try:
    users = app.state.mongo_db["users"]
    await users.create_index("email", unique=True)
    await users.create_index("username", unique=True)
    await users.create_index("status")
    print("[MongoDB] ✓ Indexes ensured on 'users' collection.", flush=True)
  except Exception as e:
    print(f"[MongoDB] ⚠  Could not create indexes: {e}", flush=True)

  try:
    await app.state.mongo_db[CLASSES_COLLECTION].create_index("name")
    await app.state.mongo_db[QUEUE_COLLECTION].create_index("created_at")
    await app.state.mongo_db[QUEUE_COLLECTION].create_index("status")
    await _migrate_classes_json_to_mongo_if_needed(app.state.mongo_db)
    await _migrate_queue_json_to_mongo_if_needed(app.state.mongo_db)
    print("[MongoDB] ✓ Queue/classes collections are ready.", flush=True)
  except Exception as e:
    print(f"[MongoDB] ⚠  Queue/classes setup failed: {e}", flush=True)

  admin_email_norm = _admin_email_norm()
  try:
    existing = await users.find_one({"email": admin_email_norm})
    if existing is not None:
      await users.update_one(
        {"_id": existing.get("_id")},
        {"$set": {"is_admin": True, "role": "admin", "status": "approved", "updated_at": _utc_now_iso()}},
      )
      print(f"[MongoDB] ✓ Admin '{admin_email_norm}' marked approved+admin.", flush=True)
    else:
      print(f"[MongoDB] ℹ  Admin '{admin_email_norm}' not registered yet — auto-approved on first register.", flush=True)
  except Exception as e:
    print(f"[MongoDB] ⚠  Admin bootstrap error: {e}", flush=True)



@app.on_event("shutdown")
async def _shutdown_mongo_after_env_loaded():
  client = getattr(app.state, "mongo_client", None)
  try:
    if client is not None:
      client.close()
  except Exception:
    pass

STORAGE_DIR = BASE_DIR / "storage"
UPLOADS_DIR = STORAGE_DIR / "uploads"
DATASETS_DIR = STORAGE_DIR / "datasets"
RUNS_DIR = STORAGE_DIR / "runs"

CLASSES_PATH = STORAGE_DIR / "classes.json"
QUEUE_PATH = STORAGE_DIR / "queue.json"

UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
DATASETS_DIR.mkdir(parents=True, exist_ok=True)
RUNS_DIR.mkdir(parents=True, exist_ok=True)


api.mount("/files", StaticFiles(directory=str(STORAGE_DIR)), name="files")


@api.get("/health")
def health():
  return {"status": "ok"}


@api.get("/mongodb/health")
async def mongodb_health(db=Depends(_get_db)):
  try:
    # Ping the database to check connectivity
    await db.command("ping")
    return {"status": "online", "database": "connected"}
  except Exception as e:
    raise HTTPException(status_code=503, detail=f"MONGODB_OFFLINE: {str(e)}")


@api.post("/auth/register")
async def auth_register(payload: dict[str, Any], db=Depends(_get_db)):
  email_raw = payload.get("email")
  username_raw = payload.get("username")
  password_raw = payload.get("password")

  if not isinstance(email_raw, str) or not isinstance(username_raw, str) or not isinstance(password_raw, str):
    raise HTTPException(status_code=400, detail="INVALID_PAYLOAD")

  email_raw = email_raw.strip()
  username_raw = username_raw.strip()
  password_raw = password_raw.strip()

  if len(username_raw) < 3 or len(username_raw) > 50:
    raise HTTPException(status_code=400, detail="INVALID_USERNAME")
  if not re.fullmatch(r"[a-zA-Z0-9_.-]+", username_raw):
    raise HTTPException(status_code=400, detail="INVALID_USERNAME")
  if len(password_raw) < 8:
    raise HTTPException(status_code=400, detail="WEAK_PASSWORD")

  try:
    v = validate_email(email_raw)
    email_norm = v.email.lower()
  except EmailNotValidError:
    raise HTTPException(status_code=400, detail="INVALID_EMAIL")

  admin_email_norm = _admin_email_norm()
  is_admin = bool(email_norm == admin_email_norm)
  status = "approved" if is_admin else "pending"
  role = "admin" if is_admin else "user"

  user_id = uuid.uuid4().hex
  doc = {
    "_id": user_id,
    "email": email_norm,
    "username": username_raw,
    "password_hash": _hash_password(password_raw),
    "status": status,
    "is_admin": is_admin,
    "role": role,
    "permissions": ["trainer:all"] if is_admin else [],
    "created_at": _utc_now_iso(),
    "updated_at": _utc_now_iso(),
  }

  try:
    await db["users"].insert_one(doc)
  except Exception as e:
    msg = str(e)
    if "E11000" in msg and "email" in msg:
      raise HTTPException(status_code=409, detail="EMAIL_ALREADY_EXISTS")
    if "E11000" in msg and "username" in msg:
      raise HTTPException(status_code=409, detail="USERNAME_ALREADY_EXISTS")
    raise HTTPException(status_code=500, detail="USER_CREATE_FAILED")

  return {"status": status, "user_id": user_id}


@api.post("/auth/login")
async def auth_login(payload: dict[str, Any], response: Response, req: FastAPIRequest, db=Depends(_get_db)):
  email_raw = payload.get("email")
  password_raw = payload.get("password")

  if not isinstance(email_raw, str) or not isinstance(password_raw, str):
    raise HTTPException(status_code=400, detail="INVALID_PAYLOAD")

  try:
    v = validate_email(email_raw.strip())
    email_norm = v.email.lower()
  except EmailNotValidError:
    raise HTTPException(status_code=400, detail="INVALID_EMAIL")

  user = await db["users"].find_one({"email": email_norm})
  if not user:
    raise HTTPException(status_code=401, detail="INVALID_CREDENTIALS")
  if user.get("status") != "approved":
    raise HTTPException(status_code=403, detail="USER_NOT_APPROVED")

  password_hash = user.get("password_hash")
  if not isinstance(password_hash, str) or not _verify_password(password_raw, password_hash):
    raise HTTPException(status_code=401, detail="INVALID_CREDENTIALS")

  permissions = _normalize_permissions(user.get("permissions"))
  token = _create_access_token(user_id=user.get("_id"), is_admin=bool(user.get("is_admin")), permissions=permissions)

  is_https = (req.url.scheme == "https")
  secure_cookie = _is_production() and is_https
  cookie_domain = _cookie_domain()

  response.set_cookie(
    key=_auth_cookie_name(),
    value=token,
    httponly=True,
    secure=secure_cookie,
    samesite="lax",
    path="/",
    domain=cookie_domain,
  )

  response.set_cookie(
    key="trainer_logged_in",
    value="1",
    httponly=False,
    secure=secure_cookie,
    samesite="lax",
    path="/",
    domain=cookie_domain,
  )

  await log_activity(req, "LOGIN", details=f"User logged in: {user.get('username')}", user=user, db=db)

  return {"ok": True}


@api.post("/auth/logout")
async def auth_logout(response: Response, req: FastAPIRequest, db=Depends(_get_db)):
  # Best effort logging before clearing cookies
  # We might not have the user context here if token is gone, 
  # but _get_current_user might still work if called.
  # For logout, we'll try to get the user context if possible.
  try:
    user = await _get_current_user(req, db)
    if user:
      await log_activity(req, "LOGOUT", details=f"User logged out: {user.get('username')}", user=user, db=db)
  except Exception:
    pass

  cookie_domain = _cookie_domain()
  response.delete_cookie(key=_auth_cookie_name(), path="/", domain=cookie_domain)
  response.delete_cookie(key="trainer_logged_in", path="/", domain=cookie_domain)
  return {"ok": True}


@api.get("/auth/me")
async def auth_me(user: dict[str, Any] = Depends(_get_current_user)):
  return {
    "id": user.get("_id"),
    "email": user.get("email"),
    "username": user.get("username"),
    "status": user.get("status"),
    "is_admin": bool(user.get("is_admin")),
    "role": _normalize_role(user.get("role")),
    "permissions": _normalize_permissions(user.get("permissions")),
  }


@api.get("/content-calendar")
async def content_calendar_list(
  limit: int = 200,
  skip: int = 0,
  _: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  lim = max(1, min(int(limit), 500))
  sk = max(0, int(skip))

  items: list[dict[str, Any]] = []
  cursor = (
    db["content_calendar"]
    .find({}, {"_id": 1, "fields": 1, "publish_date": 1, "created_at": 1, "updated_at": 1})
    .sort([("publish_date", 1), ("created_at", -1)])
    .skip(sk)
    .limit(lim)
  )
  async for doc in cursor:
    items.append(
      {
        "id": doc.get("_id"),
        "fields": doc.get("fields") or {},
        "publish_date": doc.get("publish_date"),
        "created_at": doc.get("created_at"),
        "updated_at": doc.get("updated_at"),
      }
    )

  return {"items": items, "limit": lim, "skip": sk}


@api.post("/content-calendar")
async def content_calendar_create(
  req: FastAPIRequest,
  payload: dict[str, Any],
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  raw_fields = payload.get("fields") if isinstance(payload, dict) else None
  fields = _sanitize_content_calendar_fields(raw_fields)

  now = _utc_now_iso()
  doc_id = str(uuid.uuid4())
  publish_date = fields.get("Publish Date")
  publish_date_norm = publish_date if isinstance(publish_date, str) and publish_date.strip() else None

  doc = {
    "_id": doc_id,
    "fields": fields,
    "publish_date": publish_date_norm,
    "created_at": now,
    "updated_at": now,
    "created_by": user.get("_id"),
    "updated_by": user.get("_id"),
  }

  await db["content_calendar"].insert_one(doc)
  await log_activity(req, "CONTENT_CALENDAR_CREATE", details=f"Created calendar item: {fields.get('Topic') or doc_id}", resource_id=doc_id, user=user, db=db)
  return {
    "id": doc_id,
    "fields": fields,
    "publish_date": publish_date_norm,
    "created_at": now,
    "updated_at": now,
  }


@api.patch("/content-calendar/{item_id}")
async def content_calendar_update(
  item_id: str,
  req: FastAPIRequest,
  payload: dict[str, Any],
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  raw_fields = payload.get("fields") if isinstance(payload, dict) else None
  patch_fields = _sanitize_content_calendar_fields(raw_fields)
  if not patch_fields:
    raise HTTPException(status_code=400, detail="EMPTY_PATCH")

  now = _utc_now_iso()
  update_doc: dict[str, Any] = {"updated_at": now, "updated_by": user.get("_id")}
  for k, v in patch_fields.items():
    update_doc[f"fields.{k}"] = v
  if "Publish Date" in patch_fields:
    v = patch_fields.get("Publish Date")
    update_doc["publish_date"] = v if isinstance(v, str) and v.strip() else None

  res = await db["content_calendar"].update_one({"_id": item_id}, {"$set": update_doc})
  if res.matched_count == 0:
    raise HTTPException(status_code=404, detail="NOT_FOUND")

  await log_activity(req, "CONTENT_CALENDAR_EDIT", details=f"Updated calendar item: {item_id}. Fields: {', '.join(patch_fields.keys())}", resource_id=item_id, user=user, db=db)

  doc = await db["content_calendar"].find_one(
    {"_id": item_id},
    {"_id": 1, "fields": 1, "publish_date": 1, "created_at": 1, "updated_at": 1},
  )
  if not doc:
    raise HTTPException(status_code=404, detail="NOT_FOUND")

  return {
    "id": doc.get("_id"),
    "fields": doc.get("fields") or {},
    "publish_date": doc.get("publish_date"),
    "created_at": doc.get("created_at"),
    "updated_at": doc.get("updated_at"),
  }


@api.delete("/content-calendar/{item_id}")
async def content_calendar_delete(
  item_id: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  res = await db["content_calendar"].delete_one({"_id": item_id})
  if res.deleted_count == 0:
    raise HTTPException(status_code=404, detail="NOT_FOUND")
  
  await log_activity(req, "CONTENT_CALENDAR_DELETE", details=f"Deleted calendar item: {item_id}", resource_id=item_id, user=user, db=db)
  return {"ok": True}


@api.get("/content-calendar/field-options")
async def content_calendar_list_field_options(
  _: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  return {"options": await _get_all_calendar_field_options(db)}


@api.post("/content-calendar/field-options/add")
async def content_calendar_add_field_option(
  req: FastAPIRequest,
  payload: dict[str, Any],
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  field_name = payload.get("field") if isinstance(payload, dict) else None
  raw_value = payload.get("value") if isinstance(payload, dict) else None
  if not isinstance(field_name, str) or field_name not in CONTENT_CALENDAR_SELECTABLE_FIELDS:
    raise HTTPException(status_code=400, detail="INVALID_FIELD")
  if not isinstance(raw_value, str) or not raw_value.strip():
    raise HTTPException(status_code=400, detail="VALUE_REQUIRED")

  value = raw_value.strip()
  options = await _add_calendar_field_option(db, field_name, value)
  all_options = await _get_all_calendar_field_options(db)

  await log_activity(
    req,
    "CONTENT_CALENDAR_FIELD_OPTION_ADD",
    details=f'Added "{field_name}" option "{value}"',
    resource_id=f"{field_name}:{value}",
    user=user,
    db=db,
  )

  return {"ok": True, "field": field_name, "value": value, "options": options, "all_options": all_options}


@api.post("/content-calendar/field-options/remove")
async def content_calendar_remove_field_option(
  req: FastAPIRequest,
  payload: dict[str, Any],
  admin: dict[str, Any] = Depends(_require_admin),
  db=Depends(_get_db),
):
  field_name = payload.get("field") if isinstance(payload, dict) else None
  raw_value = payload.get("value") if isinstance(payload, dict) else None
  if not isinstance(field_name, str) or field_name not in CONTENT_CALENDAR_SELECTABLE_FIELDS:
    raise HTTPException(status_code=400, detail="INVALID_FIELD")
  if not isinstance(raw_value, str) or not raw_value.strip():
    raise HTTPException(status_code=400, detail="VALUE_REQUIRED")

  value = raw_value.strip()
  updated_items = await _remove_calendar_field_option_from_all_cells(
    db,
    field_name,
    value,
    admin.get("_id"),
  )
  options = await _remove_calendar_field_option_from_registry(db, field_name, value)
  all_options = await _get_all_calendar_field_options(db)

  await log_activity(
    req,
    "CONTENT_CALENDAR_FIELD_OPTION_REMOVE",
    details=f'Removed "{field_name}" option "{value}" from {updated_items} item(s)',
    resource_id=f"{field_name}:{value}",
    user=admin,
    db=db,
  )

  return {
    "ok": True,
    "field": field_name,
    "value": value,
    "updated_items": updated_items,
    "options": options,
    "all_options": all_options,
  }


_CAMPAIGN_COLOR_PRESETS = {
  "#6366f1",
  "#8b5cf6",
  "#ec4899",
  "#f43f5e",
  "#f59e0b",
  "#10b981",
  "#06b6d4",
  "#3b82f6",
}


def _normalize_campaign_color(raw: Any) -> str:
  if not isinstance(raw, str):
    return "#6366f1"
  v = raw.strip()
  if not v.startswith("#"):
    v = f"#{v}"
  if len(v) == 7 and v.lower() in _CAMPAIGN_COLOR_PRESETS:
    return v.lower()
  if len(v) == 7 and all(c in "0123456789abcdef#" for c in v.lower()):
    return v.lower()
  return "#6366f1"


def _normalize_iso_date(raw: Any) -> str | None:
  if not isinstance(raw, str):
    return None
  s = raw.strip()
  if not s:
    return None
  if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
    return s
  return None


def _sanitize_campaign_payload(raw: Any) -> dict[str, Any]:
  if not isinstance(raw, dict):
    raise HTTPException(status_code=400, detail="INVALID_PAYLOAD")

  name = raw.get("name")
  if not isinstance(name, str) or not name.strip():
    raise HTTPException(status_code=400, detail="NAME_REQUIRED")

  start_date = _normalize_iso_date(raw.get("start_date"))
  if not start_date:
    raise HTTPException(status_code=400, detail="START_DATE_REQUIRED")

  end_date_raw = raw.get("end_date")
  end_date = _normalize_iso_date(end_date_raw) if end_date_raw not in (None, "") else None
  if end_date and end_date < start_date:
    raise HTTPException(status_code=400, detail="END_DATE_BEFORE_START")

  goal = raw.get("goal")
  goal_out = goal.strip() if isinstance(goal, str) else ""

  channels_raw = raw.get("channels")
  channels_out = ""
  if isinstance(channels_raw, str):
    channels_out = channels_raw.strip()
  elif isinstance(channels_raw, list):
    channels_out = ", ".join(str(c).strip() for c in channels_raw if str(c).strip())

  is_critical = raw.get("is_critical")
  if isinstance(is_critical, bool):
    is_critical_out = is_critical
  elif isinstance(is_critical, str):
    is_critical_out = is_critical.strip().lower() in ("1", "true", "yes", "on")
  else:
    is_critical_out = bool(is_critical)

  return {
    "name": name.strip(),
    "start_date": start_date,
    "end_date": end_date,
    "color": _normalize_campaign_color(raw.get("color")),
    "goal": goal_out,
    "channels": channels_out,
    "is_critical": is_critical_out,
  }


def _campaign_to_response(doc: dict[str, Any]) -> dict[str, Any]:
  start_date = doc.get("start_date")
  end_date = doc.get("end_date")
  effective_end = end_date if isinstance(end_date, str) and end_date.strip() else start_date
  return {
    "id": doc.get("_id"),
    "name": doc.get("name") or "",
    "start_date": start_date,
    "end_date": end_date,
    "effective_end_date": effective_end,
    "color": doc.get("color") or "#6366f1",
    "goal": doc.get("goal") or "",
    "channels": doc.get("channels") or "",
    "is_critical": bool(doc.get("is_critical")),
    "created_at": doc.get("created_at"),
    "updated_at": doc.get("updated_at"),
  }


@api.get("/marketing-campaigns")
async def marketing_campaigns_list(
  limit: int = 200,
  skip: int = 0,
  _: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  lim = max(1, min(int(limit), 500))
  sk = max(0, int(skip))

  items: list[dict[str, Any]] = []
  cursor = (
    db["marketing_campaigns"]
    .find({})
    .sort([("start_date", -1), ("created_at", -1)])
    .skip(sk)
    .limit(lim)
  )
  async for doc in cursor:
    items.append(_campaign_to_response(doc))

  return {"items": items, "limit": lim, "skip": sk}


@api.post("/marketing-campaigns")
async def marketing_campaigns_create(
  req: FastAPIRequest,
  payload: dict[str, Any],
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  data = _sanitize_campaign_payload(payload)
  now = _utc_now_iso()
  doc_id = str(uuid.uuid4())

  doc = {
    "_id": doc_id,
    **data,
    "created_at": now,
    "updated_at": now,
    "created_by": user.get("_id"),
    "updated_by": user.get("_id"),
  }

  await db["marketing_campaigns"].insert_one(doc)
  await log_activity(
    req,
    "MARKETING_CAMPAIGN_CREATE",
    details=f"Created campaign: {data.get('name')}",
    resource_id=doc_id,
    user=user,
    db=db,
  )
  return _campaign_to_response(doc)


@api.patch("/marketing-campaigns/{campaign_id}")
async def marketing_campaigns_update(
  campaign_id: str,
  req: FastAPIRequest,
  payload: dict[str, Any],
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  data = _sanitize_campaign_payload(payload)
  now = _utc_now_iso()
  update_doc: dict[str, Any] = {**data, "updated_at": now, "updated_by": user.get("_id")}

  res = await db["marketing_campaigns"].update_one({"_id": campaign_id}, {"$set": update_doc})
  if res.matched_count == 0:
    raise HTTPException(status_code=404, detail="NOT_FOUND")

  await log_activity(
    req,
    "MARKETING_CAMPAIGN_EDIT",
    details=f"Updated campaign: {campaign_id}",
    resource_id=campaign_id,
    user=user,
    db=db,
  )

  doc = await db["marketing_campaigns"].find_one({"_id": campaign_id})
  if not doc:
    raise HTTPException(status_code=404, detail="NOT_FOUND")
  return _campaign_to_response(doc)


@api.delete("/marketing-campaigns/{campaign_id}")
async def marketing_campaigns_delete(
  campaign_id: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_get_current_user),
  db=Depends(_get_db),
):
  res = await db["marketing_campaigns"].delete_one({"_id": campaign_id})
  if res.deleted_count == 0:
    raise HTTPException(status_code=404, detail="NOT_FOUND")

  await log_activity(
    req,
    "MARKETING_CAMPAIGN_DELETE",
    details=f"Deleted campaign: {campaign_id}",
    resource_id=campaign_id,
    user=user,
    db=db,
  )
  return {"ok": True}


@api.get("/admin/users")
async def admin_users(_: dict[str, Any] = Depends(_require_admin), db=Depends(_get_db)):
  out: list[dict[str, Any]] = []
  async for u in db["users"].find({}, {"password_hash": 0}).sort("created_at", -1):
    out.append(
      {
        "id": u.get("_id"),
        "email": u.get("email"),
        "username": u.get("username"),
        "status": u.get("status"),
        "is_admin": bool(u.get("is_admin")),
        "role": _normalize_role(u.get("role")),
        "permissions": _normalize_permissions(u.get("permissions")),
        "created_at": u.get("created_at"),
        "updated_at": u.get("updated_at"),
      }
    )
  return {"users": out}


@api.patch("/admin/users/{user_id}")
async def admin_update_user(
  user_id: str,
  payload: dict[str, Any],
  req: FastAPIRequest,
  admin_user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  patch: dict[str, Any] = {}

  if "status" in payload:
    status = payload.get("status")
    if status not in ["pending", "approved", "disabled"]:
      raise HTTPException(status_code=400, detail="INVALID_STATUS")
    patch["status"] = status

  if "is_admin" in payload:
    is_admin = payload.get("is_admin")
    if not isinstance(is_admin, bool):
      raise HTTPException(status_code=400, detail="INVALID_IS_ADMIN")
    patch["is_admin"] = is_admin

  if "permissions" in payload:
    patch["permissions"] = _normalize_permissions(payload.get("permissions"))

  if "role" in payload:
    patch["role"] = _normalize_role(payload.get("role"))

  if not patch:
    raise HTTPException(status_code=400, detail="EMPTY_PATCH")

  patch["updated_at"] = _utc_now_iso()
  res = await db["users"].update_one({"_id": user_id}, {"$set": patch})
  if res.matched_count == 0:
    raise HTTPException(status_code=404, detail="USER_NOT_FOUND")

  # Log the update
  updated_user = await db["users"].find_one({"_id": user_id})
  if updated_user:
    await log_activity(
      req=req,
      action="USER_UPDATE", 
      details=f"Updated user: {updated_user.get('username')} ({user_id})",
      resource_id=user_id,
      user=admin_user,
      db=db
    )

  return {"ok": True}

  return {"ok": True}


@api.get("/admin/activity-logs")
async def admin_activity_logs(
  limit: int = 100,
  skip: int = 0,
  search: str | None = None,
  action: str | None = None,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db)
):
  query: dict[str, Any] = {}
  if action:
    query["action"] = action
  if search:
    query["$or"] = [
      {"username": {"$regex": search, "$options": "i"}},
      {"details": {"$regex": search, "$options": "i"}},
      {"resource_id": {"$regex": search, "$options": "i"}},
      {"ip_address": {"$regex": search, "$options": "i"}},
      {"device": {"$regex": search, "$options": "i"}},
    ]

  cursor = db["activity_logs"].find(query).sort("timestamp", -1).skip(skip).limit(limit)
  logs = []
  async for doc in cursor:
    doc["id"] = str(doc.pop("_id"))
    logs.append(doc)

  total = await db["activity_logs"].count_documents(query)
  return {"logs": logs, "total": total}


@api.get("/admin/products/field-changes")
async def admin_product_field_changes(
  limit: int = 5000,
  record_ids: str | None = None,
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  """Return per-product, per-field edit history for admin moderation UI."""
  limit = max(1, min(limit, 10000))
  query: dict[str, Any] = {
    "action": {"$in": ["PRODUCT_EDIT", "PRODUCT_INLINE_EDIT", "PRODUCT_IMPORT_EDIT"]},
  }
  if record_ids:
    ids = [part.strip() for part in record_ids.split(",") if part.strip()]
    if ids:
      query["resource_id"] = {"$in": ids}

  cursor = db["activity_logs"].find(query).sort("timestamp", -1).limit(limit)
  changes: dict[str, dict[str, list[dict[str, Any]]]] = {}

  async for doc in cursor:
    rid = doc.get("resource_id")
    if not rid:
      continue
    rid = str(rid)
    if rid not in changes:
      changes[rid] = {}

    base_entry = {
      "id": str(doc.get("_id")),
      "username": doc.get("username") or "unknown",
      "timestamp": doc.get("timestamp"),
      "action": doc.get("action") or "PRODUCT_EDIT",
    }

    structured = doc.get("field_changes")
    if isinstance(structured, list) and structured:
      for fc in structured:
        if not isinstance(fc, dict):
          continue
        field_name = str(fc.get("field") or "").strip()
        if not field_name:
          continue
        old_value = _serialize_field_value(fc.get("old_value"))
        new_value = _serialize_field_value(fc.get("new_value"))
        change_type = fc.get("change_type") or _infer_field_change_type(old_value, new_value)
        entry = {
          **base_entry,
          "field": field_name,
          "change_type": change_type,
          "old_value": old_value,
          "new_value": new_value,
        }
        changes[rid].setdefault(field_name, []).append(entry)
      continue

    for field_name in _parse_product_edit_fields_from_details(doc.get("details") or ""):
      entry = {
        **base_entry,
        "field": field_name,
        "change_type": "update",
        "old_value": "",
        "new_value": "",
      }
      changes[rid].setdefault(field_name, []).append(entry)

  return {"changes": changes}


@api.post("/admin/log-event")
async def admin_log_event(
  payload: dict[str, Any],
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_get_current_user),
  db: Any = Depends(_get_db)
):
  action = payload.get("action")
  details = payload.get("details", "")
  resource_id = payload.get("resource_id")
  
  if not action:
    raise HTTPException(status_code=400, detail="ACTION_REQUIRED")
    
  await log_activity(req, action, details=details, resource_id=resource_id, user=user, db=db)
  return {"ok": True}


@api.get("/dam/assets")
async def dam_assets(
  _: dict[str, Any] = Depends(_get_current_user), 
  db=Depends(_get_db)
):
  records: list[dict[str, Any]] = []
  cursor = db["dam_assets"].find({}).sort("created_at", -1).limit(2000)
  
  async for doc in cursor:
    records.append({
      "id": doc.get("_id"),
      "fields": doc.get("fields") or {},
      "createdTime": doc.get("created_at")
    })

  columns_set: set[str] = set()
  for r in records:
    f = r.get("fields")
    if isinstance(f, dict):
      for k in f.keys():
        columns_set.add(k)

  # Ensure the DAM column is always present for the UI
  columns_set.add("DAM")

  columns = sorted(columns_set)
  return {"columns": columns, "records": records, "count": len(records)}


PRODUCTS_ASSETS_DEFAULT_LIMIT = 500
PRODUCTS_ASSETS_MAX_LIMIT = 1000

# Cursor payload is intentionally opaque to clients; it stores the last row boundary.
def _encode_products_assets_cursor(created_time: Any, record_id: Any) -> str:
  payload = {
    "createdTime": str(created_time or ""),
    "id": str(record_id or ""),
  }
  raw = json.dumps(payload, separators=(",", ":"), ensure_ascii=True).encode("utf-8")
  return base64.urlsafe_b64encode(raw).decode("ascii")

def _decode_products_assets_cursor(cursor: str) -> tuple[str, str]:
  try:
    padded = cursor + "=" * (-len(cursor) % 4)
    decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
    parsed = json.loads(decoded)
    created_time = str(parsed.get("createdTime") or "")
    record_id = str(parsed.get("id") or "")
    if not created_time or not record_id:
      raise ValueError("cursor fields missing")
    return created_time, record_id
  except Exception as exc:
    raise HTTPException(status_code=400, detail="INVALID_CURSOR") from exc

# Offset cursor used by the `sort=num` mode (business order by the Num column).
def _encode_offset_cursor(offset: int) -> str:
  raw = json.dumps({"offset": int(offset)}, separators=(",", ":")).encode("utf-8")
  return base64.urlsafe_b64encode(raw).decode("ascii")

def _decode_offset_cursor(cursor: str) -> int:
  try:
    padded = cursor + "=" * (-len(cursor) % 4)
    decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
    parsed = json.loads(decoded)
    return max(0, int(parsed.get("offset") or 0))
  except Exception as exc:
    raise HTTPException(status_code=400, detail="INVALID_CURSOR") from exc

def _build_products_assets_columns(records: list[dict[str, Any]]) -> list[str]:
  columns_set: set[str] = set()
  for r in records:
    f = r.get("fields")
    if isinstance(f, dict):
      for k in f.keys():
        columns_set.add(k)
  columns_set.add("DAM")
  columns_set.add("Space")
  columns_set.add("Color")
  columns_set.add("Material")
  columns_set.add("Content Status")
  return sorted(columns_set)

async def _products_assets_page(
  db: Any,
  limit: int,
  cursor: str | None,
  sort: str = "recent",
) -> dict[str, Any]:
  safe_limit = max(1, min(limit, PRODUCTS_ASSETS_MAX_LIMIT))

  # Business order: sort by the numeric `Num` column (ascending). Non-numeric /
  # missing Num values sort last. Offset-paginated so progressive pages append in
  # a stable order (page 1 = lowest 40 Num, page 2 = next, …) — no client reorder.
  if sort == "num":
    offset = _decode_offset_cursor(cursor) if cursor else 0
    big = 1e18
    pipeline = [
      {"$addFields": {
        "_numKey": {"$convert": {"input": "$fields.Num", "to": "double", "onError": big, "onNull": big}},
      }},
      {"$sort": {"_numKey": 1, "_id": 1}},
      {"$skip": offset},
      {"$limit": safe_limit + 1},
    ]
    docs = await db["products"].aggregate(pipeline, allowDiskUse=True).to_list(length=safe_limit + 1)
    has_more = len(docs) > safe_limit
    page_docs = docs[:safe_limit]
    records = [
      {"id": d.get("_id"), "fields": d.get("fields") or {}, "createdTime": d.get("created_at")}
      for d in page_docs
    ]
    next_cursor = _encode_offset_cursor(offset + safe_limit) if has_more else None
    columns = _build_products_assets_columns(records)
    return {
      "columns": columns,
      "records": records,
      "count": len(records),
      "has_more": has_more,
      "next_cursor": next_cursor,
      "page_size": safe_limit,
      "sort": "num",
    }

  query: dict[str, Any] = {}

  if cursor:
    cursor_created_time, cursor_id = _decode_products_assets_cursor(cursor)
    # Stable seek pagination: fetch rows strictly "older" than the cursor tuple
    # based on the same compound sort used below (created_at desc, _id desc).
    query = {
      "$or": [
        {"created_at": {"$lt": cursor_created_time}},
        {
          "created_at": cursor_created_time,
          "_id": {"$lt": cursor_id},
        },
      ]
    }

  # Read one extra row to compute has_more/next_cursor without a second query.
  docs = await db["products"].find(query).sort([("created_at", -1), ("_id", -1)]).limit(safe_limit + 1).to_list(
    length=safe_limit + 1
  )

  has_more = len(docs) > safe_limit
  page_docs = docs[:safe_limit]

  records = [
    {
      "id": doc.get("_id"),
      "fields": doc.get("fields") or {},
      "createdTime": doc.get("created_at"),
    }
    for doc in page_docs
  ]

  next_cursor = None
  if has_more and page_docs:
    # Next cursor points to the last returned row, never to the lookahead row.
    last = page_docs[-1]
    next_cursor = _encode_products_assets_cursor(last.get("created_at"), last.get("_id"))

  columns = _build_products_assets_columns(records)
  return {
    "columns": columns,
    "records": records,
    "count": len(records),
    "has_more": has_more,
    "next_cursor": next_cursor,
    "page_size": safe_limit,
  }

@api.get("/public/products/assets")
async def public_products_assets(
  limit: int = Query(PRODUCTS_ASSETS_DEFAULT_LIMIT, ge=1, le=PRODUCTS_ASSETS_MAX_LIMIT),
  cursor: str | None = Query(None),
  sort: str = Query("recent"),
  db=Depends(_get_db),
):
  sort_mode = "num" if str(sort).strip().lower() == "num" else "recent"
  return await _products_assets_page(db, limit=limit, cursor=cursor, sort=sort_mode)


@api.get("/public/products/field-options")
async def public_product_field_options(db=Depends(_get_db)):
  return {"options": await _get_product_field_options(db)}


_PRODUCT_IMAGE_FIELD_RE = re.compile(r"^Image(\d+)?$", re.IGNORECASE)


def _extract_http_urls_from_field_value(value: Any) -> list[str]:
  urls: list[str] = []
  if value is None:
    return urls
  if isinstance(value, str):
    text = value.strip()
    if not text:
      return urls
    if text.startswith("http://") or text.startswith("https://"):
      urls.append(text)
      return urls
    for part in re.split(r"[\n,]+", text):
      candidate = part.strip()
      if candidate.startswith("http://") or candidate.startswith("https://"):
        urls.append(candidate)
    return urls
  if isinstance(value, list):
    for item in value:
      urls.extend(_extract_http_urls_from_field_value(item))
    return urls
  if isinstance(value, dict):
    for nested in value.values():
      urls.extend(_extract_http_urls_from_field_value(nested))
  return urls


async def _count_product_image_urls(db: Any) -> int:
  total = 0
  cursor = db["products"].find({}, {"fields": 1})
  async for doc in cursor:
    fields = doc.get("fields") or {}
    for key, value in fields.items():
      if key == "Image" or _PRODUCT_IMAGE_FIELD_RE.fullmatch(str(key)):
        total += len(_extract_http_urls_from_field_value(value))
  return total


async def _count_distinct_product_categories(db: Any) -> int:
  labels: set[str] = set()
  cursor = db["products"].find({}, {"fields.Category": 1})
  async for doc in cursor:
    raw = (doc.get("fields") or {}).get("Category")
    if not isinstance(raw, str):
      continue
    for part in raw.split(","):
      label = part.strip()
      if label:
        labels.add(label)
  return len(labels)


@api.get("/public/platform/stats")
async def public_platform_stats(db=Depends(_get_db)):
  products_count, product_images_count, dam_count, users_count, calendar_count, yolo_classes_count, category_labels_count = await asyncio.gather(
    db["products"].count_documents({}),
    _count_product_image_urls(db),
    db["dam_assets"].count_documents({}),
    db["users"].count_documents({}),
    db["content_calendar"].count_documents({}),
    db[CLASSES_COLLECTION].count_documents({}),
    _count_distinct_product_categories(db),
  )

  db_status = "connected"
  try:
    await db.client.admin.command("ping")
  except Exception:
    db_status = "disconnected"

  return {
    "products_count": products_count,
    "product_images_count": product_images_count,
    "dam_assets_count": dam_count,
    "users_count": users_count,
    "calendar_items_count": calendar_count,
    "yolo_classes_count": yolo_classes_count,
    "category_labels_count": category_labels_count,
    "db_status": db_status,
    "updated_at": datetime.now(timezone.utc).isoformat(),
  }



_SALES_RESTRICTED_PRODUCT_FIELDS = {
  "code number",
  "code no",
  "code",
  "collection name",
  "colecction name",
  "name",
  "price",
  "collection code",
  "colecction code",
  "variant number",
  "num",
  "factory code",
}


def _is_sales_restricted_product_field(field_name: str) -> bool:
  return field_name.strip().lower() in _SALES_RESTRICTED_PRODUCT_FIELDS


@api.patch("/products/assets/{record_id}")
async def patch_product_asset(
  record_id: str,
  payload: dict[str, Any],
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_get_current_user),
  db: Any = Depends(_get_db),
):
  role = _normalize_role(user.get("role"))
  if role not in ["admin", "sales"]:
    raise HTTPException(status_code=403, detail="FORBIDDEN_ROLE")

  fields_to_update = payload.get("fields")
  if not isinstance(fields_to_update, dict):
    raise HTTPException(status_code=400, detail="INVALID_PAYLOAD_EXPECTED_FIELDS")

  is_platform_admin = user.get("is_admin") is True or role == "admin"
  if role == "sales" and not is_platform_admin:
    for field_key in fields_to_update.keys():
      if _is_sales_restricted_product_field(str(field_key)):
        raise HTTPException(
          status_code=403,
          detail=f"FIELD_NOT_EDITABLE_BY_SALES:{field_key}",
        )

  doc_before = await db["products"].find_one({"_id": record_id})
  if not doc_before:
    raise HTTPException(status_code=404, detail="PRODUCT_NOT_FOUND")

  old_fields = doc_before.get("fields") or {}
  field_changes: list[dict[str, Any]] = []
  for k, v in fields_to_update.items():
    old_v = old_fields.get(k)
    if old_v is None:
      for ok, ov in old_fields.items():
        if ok.strip().lower() == k.strip().lower():
          old_v = ov
          break
    old_s = _serialize_field_value(old_v)
    new_s = _serialize_field_value(v)
    if old_s == new_s:
      continue
    field_changes.append(
      {
        "field": k,
        "old_value": old_s,
        "new_value": new_s,
        "change_type": _infer_field_change_type(old_s, new_s),
      }
    )

  update_doc = {"updated_at": _utc_now_iso()}
  for k, v in fields_to_update.items():
    update_doc[f"fields.{k}"] = v

  res = await db["products"].update_one({"_id": record_id}, {"$set": update_doc})
  if res.matched_count == 0:
    raise HTTPException(status_code=404, detail="PRODUCT_NOT_FOUND")

  doc = await db["products"].find_one({"_id": record_id})
  
  # Log activity
  fields = doc.get("fields") or {}
  item_label = fields.get("Colecction Name") or fields.get("Name") or record_id
  await log_activity(
    req=req,
    action="PRODUCT_EDIT",
    details=f"Edited product: {item_label}. Fields: {', '.join(fields_to_update.keys())}",
    resource_id=record_id,
    user=user,
    db=db,
    field_changes=field_changes or None,
  )

  return {"id": doc["_id"], "fields": doc["fields"]}


def _product_has_identifier(fields: dict[str, Any]) -> bool:
  name_keys = {"collection name", "colecction name", "name"}
  code_keys = {"code number", "code no", "code", "factory code", "collection code", "colecction code"}
  for key, value in fields.items():
    if not _serialize_field_value(value):
      continue
    normalized = str(key).strip().lower()
    if normalized in name_keys or normalized in code_keys:
      return True
  return False


@api.post("/products/assets")
async def create_product_asset(
  payload: dict[str, Any],
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  raw_fields = payload.get("fields")
  if not isinstance(raw_fields, dict):
    raise HTTPException(status_code=400, detail="INVALID_PAYLOAD_EXPECTED_FIELDS")

  cleaned: dict[str, Any] = {}
  for key, value in raw_fields.items():
    if not isinstance(key, str) or "." in key or key.startswith("$"):
      continue
    if value is None:
      continue
    if isinstance(value, str) and not value.strip():
      continue
    cleaned[key] = value

  if not cleaned:
    raise HTTPException(status_code=400, detail="NO_FIELDS")
  if not _product_has_identifier(cleaned):
    raise HTTPException(status_code=400, detail="REQUIRES_COLLECTION_OR_CODE")

  product_id = uuid.uuid4().hex
  now = _utc_now_iso()
  doc = {
    "_id": product_id,
    "fields": cleaned,
    "created_at": now,
    "updated_at": now,
    "source": "manual",
  }
  await db["products"].insert_one(doc)

  item_label = (
    cleaned.get("Collection Name")
    or cleaned.get("Colecction Name")
    or cleaned.get("Name")
    or cleaned.get("CODE NUMBER")
    or cleaned.get("Code Number")
    or product_id
  )
  field_changes = _build_field_change_entries_from_create(cleaned)
  await log_activity(
    req=req,
    action="PRODUCT_CREATE",
    details=f"Created product: {item_label}",
    resource_id=product_id,
    user=user,
    db=db,
    field_changes=field_changes or None,
  )

  return {"id": product_id, "fields": cleaned, "createdTime": now}


@api.delete("/products/assets/{record_id}")
async def delete_product_asset(
  record_id: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  doc = await db["products"].find_one({"_id": record_id})
  if not doc:
    raise HTTPException(status_code=404, detail="PRODUCT_NOT_FOUND")

  res = await db["products"].delete_one({"_id": record_id})
  if res.deleted_count == 0:
    raise HTTPException(status_code=404, detail="PRODUCT_NOT_FOUND")

  fields = doc.get("fields") or {}
  item_label = fields.get("Colecction Name") or fields.get("Name") or record_id
  await log_activity(
    req=req,
    action="PRODUCT_DELETE",
    details=f"Deleted product: {item_label}",
    resource_id=record_id,
    user=user,
    db=db,
  )

  return {"deleted": True, "id": record_id}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Admin Dashboard Endpoints
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Collections that can be backed up / restored
_BACKUPABLE_COLLECTIONS = [
  "products",
  "dam_assets",
  "users",
  "content_calendar",
  "marketing_campaigns",
  "activity_logs",
  CONTENT_CALENDAR_FIELD_OPTIONS_COLLECTION,
  PRODUCT_FIELD_OPTIONS_COLLECTION,
  PRODUCT_IMPORT_BATCHES_COLLECTION,
  PRODUCT_IMPORT_ROWS_COLLECTION,
]


@api.get("/admin/dashboard/stats")
async def admin_dashboard_stats(
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  """Return aggregated statistics for the admin dashboard."""

  # Count documents in key collections
  products_count = await db["products"].count_documents({})
  dam_count = await db["dam_assets"].count_documents({})
  users_count = await db["users"].count_documents({})
  logs_count = await db["activity_logs"].count_documents({})
  calendar_count = await db["content_calendar"].count_documents({})

  # Category breakdown from products
  category_counts: dict[str, int] = {}
  color_counts: dict[str, int] = {}
  space_counts: dict[str, int] = {}
  material_counts: dict[str, int] = {}

  cursor = db["products"].find({}, {"fields.Category": 1, "fields.Color": 1, "fields.Space": 1, "fields.Material": 1})
  async for doc in cursor:
    fields = doc.get("fields") or {}
    for field_name, counts_dict in [
      ("Category", category_counts),
      ("Color", color_counts),
      ("Space", space_counts),
      ("Material", material_counts),
    ]:
      val = fields.get(field_name)
      if isinstance(val, str) and val.strip():
        for part in val.split(","):
          p = part.strip()
          if p:
            counts_dict[p] = counts_dict.get(p, 0) + 1

  # Recent activity logs (last 10)
  recent_logs: list[dict[str, Any]] = []
  log_cursor = db["activity_logs"].find({}).sort("timestamp", -1).limit(10)
  async for doc in log_cursor:
    doc["id"] = str(doc.pop("_id"))
    recent_logs.append(doc)

  # Recent edits (last 5 product edits)
  recent_edits: list[dict[str, Any]] = []
  edit_cursor = db["activity_logs"].find({"action": "PRODUCT_EDIT"}).sort("timestamp", -1).limit(5)
  async for doc in edit_cursor:
    doc["id"] = str(doc.pop("_id"))
    recent_edits.append(doc)

  # MongoDB connection status
  db_status = "connected"
  try:
    await db.client.admin.command("ping")
  except Exception:
    db_status = "disconnected"

  return {
    "products_count": products_count,
    "dam_count": dam_count,
    "users_count": users_count,
    "logs_count": logs_count,
    "calendar_count": calendar_count,
    "category_counts": category_counts,
    "color_counts": color_counts,
    "space_counts": space_counts,
    "material_counts": material_counts,
    "recent_logs": recent_logs,
    "recent_edits": recent_edits,
    "db_status": db_status,
    "backupable_collections": _BACKUPABLE_COLLECTIONS,
  }


@api.get("/admin/products/field-options")
async def admin_get_product_field_options(
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  return {"options": await _get_product_field_options(db)}


@api.put("/admin/products/field-options")
async def admin_update_product_field_options(
  payload: dict[str, Any],
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  options = _normalize_product_field_options_payload(payload.get("options"))
  now = _utc_now_iso()

  await db[PRODUCT_FIELD_OPTIONS_COLLECTION].update_one(
    {"_id": PRODUCT_FIELD_OPTIONS_DOC_ID},
    {
      "$set": {
        "options": options,
        "updated_at": now,
        "updated_by": user.get("_id"),
      },
      "$setOnInsert": {"created_at": now},
    },
    upsert=True,
  )

  await log_activity(
    req,
    "PRODUCT_FIELD_OPTIONS_UPDATE",
    details="Updated selectable fields: Category, Space, Color, Material",
    user=user,
    db=db,
  )

  return {"options": options, "updated_at": now}


@api.get("/admin/products/imports")
async def admin_list_product_imports(
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  imports: list[dict[str, Any]] = []
  cursor = db[PRODUCT_IMPORT_BATCHES_COLLECTION].find({}).sort("created_at", -1).limit(50)
  async for doc in cursor:
    doc["id"] = str(doc.pop("_id"))
    imports.append(doc)
  return {"imports": imports}


@api.post("/admin/products/imports/upload")
async def admin_upload_product_import(
  req: FastAPIRequest,
  file: UploadFile = File(...),
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  filename = file.filename or "products-import.xlsx"
  if not filename.lower().endswith((".xlsx", ".xlsm", ".numbers")):
    raise HTTPException(status_code=400, detail="ONLY_XLSX_XLSM_OR_NUMBERS_SUPPORTED")

  content = await file.read()
  if not content:
    raise HTTPException(status_code=400, detail="EMPTY_FILE")

  import_id = uuid.uuid4().hex
  parsed = _parse_product_import_workbook(content, filename, import_id)
  rows = parsed["rows"]
  if not rows:
    raise HTTPException(status_code=400, detail="NO_PRODUCT_ROWS_FOUND")

  now = _utc_now_iso()
  columns = parsed["columns"]
  warnings_count = sum(1 for row in rows if row.get("warnings"))

  for row in rows:
    row["import_id"] = import_id
    row["created_at"] = now
    row["updated_at"] = now

  batch_doc = {
    "_id": import_id,
    "filename": filename,
    "status": "staged",
    "columns": columns,
    "sheets": parsed["sheets"],
    "row_count": len(rows),
    "warnings_count": warnings_count,
    "created_at": now,
    "updated_at": now,
    "created_by": user.get("_id"),
  }

  await db[PRODUCT_IMPORT_BATCHES_COLLECTION].insert_one(batch_doc)
  await db[PRODUCT_IMPORT_ROWS_COLLECTION].insert_many(rows)

  await log_activity(
    req,
    "PRODUCT_IMPORT_UPLOAD",
    details=f"Uploaded product import: {filename}. Rows: {len(rows)}",
    resource_id=import_id,
    user=user,
    db=db,
  )

  return {
    "import": {**batch_doc, "id": import_id},
    "preview": [
      {
        "id": row["_id"],
        "fields": row["fields"],
        "warnings": row["warnings"],
        "source_sheet": row["source_sheet"],
        "source_row_number": row["source_row_number"],
      }
      for row in rows[:25]
    ],
  }


@api.get("/admin/products/imports/{import_id}/rows")
async def admin_get_product_import_rows(
  import_id: str,
  skip: int = 0,
  limit: int = 5000,
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  batch = await db[PRODUCT_IMPORT_BATCHES_COLLECTION].find_one({"_id": import_id})
  if not batch:
    raise HTTPException(status_code=404, detail="IMPORT_NOT_FOUND")

  safe_limit = max(1, min(limit, 5000))
  safe_skip = max(0, skip)
  total = await db[PRODUCT_IMPORT_ROWS_COLLECTION].count_documents({"import_id": import_id})
  records: list[dict[str, Any]] = []
  cursor = (
    db[PRODUCT_IMPORT_ROWS_COLLECTION]
    .find({"import_id": import_id})
    .sort([("source_sheet", 1), ("source_row_number", 1)])
    .skip(safe_skip)
    .limit(safe_limit)
  )
  async for doc in cursor:
    fields = doc.get("fields") or {}
    if isinstance(fields, dict):
      fields = {"Row": doc.get("source_row_number"), **fields}
    else:
      fields = {"Row": doc.get("source_row_number")}
    records.append({
      "id": str(doc.get("_id")),
      "fields": fields,
      "raw_fields": doc.get("raw_fields") or {},
      "warnings": doc.get("warnings") or [],
      "status": doc.get("status") or "staged",
      "source_sheet": doc.get("source_sheet"),
      "source_row_number": doc.get("source_row_number"),
    })

  batch["id"] = str(batch.pop("_id"))
  columns = batch.get("columns") or []
  if "Row" not in columns:
    columns = ["Row", *columns]
  return {
    "import": batch,
    "columns": columns,
    "records": records,
    "count": total,
    "skip": safe_skip,
    "limit": safe_limit,
  }


@api.post("/admin/products/imports/{import_id}/reprocess")
async def admin_reprocess_product_import(
  import_id: str,
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  batch = await db[PRODUCT_IMPORT_BATCHES_COLLECTION].find_one({"_id": import_id})
  if not batch:
    raise HTTPException(status_code=404, detail="IMPORT_NOT_FOUND")

  now = _utc_now_iso()
  processed_count = 0
  changed_count = 0
  cursor = db[PRODUCT_IMPORT_ROWS_COLLECTION].find({"import_id": import_id})

  async for row in cursor:
    row_id = str(row.get("_id"))
    fields = row.get("fields") or {}
    if not isinstance(fields, dict):
      continue

    next_fields = _apply_import_row_formulas(fields)
    processed_count += 1
    if next_fields == fields:
      continue

    changed_count += 1
    await db[PRODUCT_IMPORT_ROWS_COLLECTION].update_one(
      {"_id": row_id, "import_id": import_id},
      {"$set": {"fields": next_fields, "updated_at": now}},
    )

  await db[PRODUCT_IMPORT_BATCHES_COLLECTION].update_one(
    {"_id": import_id},
    {"$set": {"updated_at": now, "last_reprocessed_at": now}},
  )

  return {
    "ok": True,
    "processed_count": processed_count,
    "changed_count": changed_count,
  }


@api.patch("/admin/products/imports/{import_id}/rows/{row_id}")
async def admin_update_product_import_row(
  import_id: str,
  row_id: str,
  payload: dict[str, Any],
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  field_name = payload.get("field")
  if not isinstance(field_name, str) or not field_name.strip():
    raise HTTPException(status_code=400, detail="INVALID_FIELD")

  field_name = field_name.strip()
  if "." in field_name or field_name.startswith("$"):
    raise HTTPException(status_code=400, detail="INVALID_FIELD")

  row = await db[PRODUCT_IMPORT_ROWS_COLLECTION].find_one({"_id": row_id, "import_id": import_id})
  if not row:
    raise HTTPException(status_code=404, detail="IMPORT_ROW_NOT_FOUND")

  fields = dict(row.get("fields") or {})
  raw_fields = dict(row.get("raw_fields") or {})
  warnings = [item for item in row.get("warnings") or [] if isinstance(item, str) and not item.startswith(f"{field_name}:")]

  cleaned_value, warning = _clean_import_field(field_name, payload.get("value"))
  if cleaned_value == "":
    fields.pop(field_name, None)
    raw_fields.pop(field_name, None)
  else:
    fields[field_name] = cleaned_value
    raw_fields[field_name] = _serialize_excel_value(payload.get("value"))
    if warning:
      warnings.append(f"{field_name}: {warning}")

  fields = _apply_import_row_formulas(fields)

  now = _utc_now_iso()
  await db[PRODUCT_IMPORT_ROWS_COLLECTION].update_one(
    {"_id": row_id, "import_id": import_id},
    {
      "$set": {
        "fields": fields,
        "raw_fields": raw_fields,
        "warnings": warnings,
        "updated_at": now,
      }
    },
  )

  await log_activity(
    req,
    "PRODUCT_IMPORT_ROW_EDIT",
    details=f"Edited staged import row field: {field_name}",
    resource_id=row_id,
    user=user,
    db=db,
  )

  return {
    "id": row_id,
    "fields": fields,
    "raw_fields": raw_fields,
    "warnings": warnings,
    "status": row.get("status") or "staged",
    "source_sheet": row.get("source_sheet"),
    "source_row_number": row.get("source_row_number"),
  }


@api.post("/admin/products/imports/{import_id}/preview-match")
async def admin_preview_product_import_match(
  import_id: str,
  payload: dict[str, Any],
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  batch = await db[PRODUCT_IMPORT_BATCHES_COLLECTION].find_one({"_id": import_id})
  if not batch:
    raise HTTPException(status_code=404, detail="IMPORT_NOT_FOUND")

  match_config = _parse_import_match_payload(payload)
  if not match_config:
    raise HTTPException(status_code=400, detail="INVALID_MATCH_PAYLOAD")
  import_column, product_column = match_config

  batch_columns = batch.get("columns") or []
  if import_column != "Row" and import_column not in batch_columns:
    raise HTTPException(status_code=400, detail="IMPORT_COLUMN_NOT_IN_BATCH")

  product_index = await _build_product_index_by_column(db, product_column)
  matched_count = 0
  unmatched_count = 0
  empty_import_value_count = 0
  matched_samples: list[dict[str, Any]] = []
  unmatched_samples: list[dict[str, Any]] = []
  empty_samples: list[dict[str, Any]] = []
  row_statuses: dict[str, str] = {}
  sample_limit = 25

  cursor = db[PRODUCT_IMPORT_ROWS_COLLECTION].find({"import_id": import_id})
  async for row in cursor:
    row_id = str(row.get("_id"))
    fields = row.get("fields") or {}
    if not isinstance(fields, dict):
      fields = {}

    status = _classify_import_row_match_status(fields, product_index, import_column)
    row_statuses[row_id] = status
    import_value = _import_row_field_value(fields, import_column)

    if status == "empty":
      empty_import_value_count += 1
      if len(empty_samples) < sample_limit:
        empty_samples.append({
          "row_id": row_id,
          "import_value": _import_value_to_text(import_value) or None,
          "reason": "empty_import_value",
          "source_sheet": row.get("source_sheet"),
          "source_row_number": row.get("source_row_number"),
        })
      continue

    if status == "matched":
      matched_count += 1
      product = product_index.get(_product_compare_value(import_value))
      if len(matched_samples) < sample_limit and product:
        product_fields = product.get("fields") or {}
        product_value = _canonical_product_field_value(product_fields, product_column)
        matched_samples.append({
          "row_id": row_id,
          "product_id": str(product.get("_id")),
          "import_value": _import_value_to_text(import_value),
          "product_value": _import_value_to_text(product_value),
          "source_sheet": row.get("source_sheet"),
          "source_row_number": row.get("source_row_number"),
        })
      continue

    unmatched_count += 1
    if len(unmatched_samples) < sample_limit:
      unmatched_samples.append({
        "row_id": row_id,
        "import_value": _import_value_to_text(import_value),
        "reason": "no_product_match",
        "source_sheet": row.get("source_sheet"),
        "source_row_number": row.get("source_row_number"),
      })

  total_rows = matched_count + unmatched_count + empty_import_value_count
  return {
    "ok": True,
    "match": {
      "import_column": import_column,
      "product_column": product_column,
    },
    "total_rows": total_rows,
    "matched_count": matched_count,
    "unmatched_count": unmatched_count,
    "empty_import_value_count": empty_import_value_count,
    "row_statuses": row_statuses,
    "matched_samples": matched_samples,
    "unmatched_samples": unmatched_samples,
    "empty_samples": empty_samples,
  }


@api.post("/admin/products/imports/{import_id}/apply")
async def admin_apply_product_import(
  import_id: str,
  req: FastAPIRequest,
  payload: dict[str, Any] | None = None,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  batch = await db[PRODUCT_IMPORT_BATCHES_COLLECTION].find_one({"_id": import_id})
  if not batch:
    raise HTTPException(status_code=404, detail="IMPORT_NOT_FOUND")

  raw_columns = (payload or {}).get("columns")
  if not isinstance(raw_columns, list) or not raw_columns:
    raise HTTPException(status_code=400, detail="NO_COLUMNS_SELECTED")
  selected_columns = {
    column.strip()
    for column in raw_columns
    if isinstance(column, str) and column.strip() and "." not in column and not column.strip().startswith("$")
  }
  if not selected_columns:
    raise HTTPException(status_code=400, detail="NO_COLUMNS_SELECTED")

  batch_columns = batch.get("columns") or []
  allowed_transfer_columns = set(batch_columns) | {"Row", "Image1"}
  unknown_transfer_columns = selected_columns - allowed_transfer_columns
  if unknown_transfer_columns:
    raise HTTPException(
      status_code=400,
      detail=f"SELECTED_COLUMNS_NOT_IN_IMPORT: {sorted(unknown_transfer_columns)}",
    )

  match_config = _parse_import_match_payload(payload)
  if match_config:
    import_match_column, product_match_column = match_config
    if import_match_column != "Row" and import_match_column not in batch_columns:
      raise HTTPException(status_code=400, detail="IMPORT_COLUMN_NOT_IN_BATCH")
    product_index = await _build_product_index_by_column(db, product_match_column)
  else:
    import_match_column = None
    product_match_column = None
    product_index = await _build_product_match_index(db)

  apply_row_groups = _parse_apply_row_groups(payload) if match_config else set(PRODUCT_IMPORT_ROW_MATCH_STATUSES)
  now = _utc_now_iso()
  created_count = 0
  updated_count = 0
  unchanged_count = 0
  skipped_count = 0
  rows_processed_by_group: dict[str, int] = {status: 0 for status in PRODUCT_IMPORT_ROW_MATCH_STATUSES}
  rows_skipped_by_group: dict[str, int] = {status: 0 for status in PRODUCT_IMPORT_ROW_MATCH_STATUSES}
  changed_cells_count = 0
  created_product_ids: list[str] = []
  updated_product_ids: list[str] = []
  product_write_ops: list[Any] = []
  row_write_ops: list[Any] = []
  pending_import_audit_logs: list[tuple[str, list[dict[str, Any]], str]] = []
  import_label = str(batch.get("filename") or import_id)

  cursor = (
    db[PRODUCT_IMPORT_ROWS_COLLECTION]
    .find({"import_id": import_id})
    .sort([("source_sheet", 1), ("source_row_number", 1)])
  )

  async for row in cursor:
    row_id = str(row.get("_id"))
    staged_fields = row.get("fields") or {}
    if not isinstance(staged_fields, dict) or not staged_fields:
      unchanged_count += 1
      continue
    product_staged_fields = dict(staged_fields)
    source_row_number = row.get("source_row_number")
    if source_row_number is not None:
      product_staged_fields["Row"] = source_row_number

    if match_config:
      row_match_status = _classify_import_row_match_status(staged_fields, product_index, import_match_column)
      if row_match_status not in apply_row_groups:
        skipped_count += 1
        rows_skipped_by_group[row_match_status] = rows_skipped_by_group.get(row_match_status, 0) + 1
        continue
      rows_processed_by_group[row_match_status] = rows_processed_by_group.get(row_match_status, 0) + 1
      existing_product = _find_matching_product_by_column(staged_fields, product_index, import_match_column)
    else:
      existing_product = _find_matching_product(staged_fields, product_index)
    if existing_product:
      product_id = str(existing_product.get("_id"))
      existing_fields = existing_product.get("fields") or {}
      if not isinstance(existing_fields, dict):
        existing_fields = {}

      changed_fields = _changed_product_fields(product_staged_fields, existing_fields, product_id, selected_columns)
      if changed_fields:
        update_doc = {f"fields.{field_name}": value for field_name, value in changed_fields.items()}
        update_doc["updated_at"] = now
        product_write_ops.append(UpdateOne({"_id": product_id}, {"$set": update_doc}))
        updated_count += 1
        changed_cells_count += len(changed_fields)
        updated_product_ids.append(product_id)
        audit_entries = _build_field_change_entries_from_patch(existing_fields, changed_fields)
        if audit_entries:
          pending_import_audit_logs.append((product_id, audit_entries, "updated"))

        merged_fields = {**existing_fields, **changed_fields}
        existing_product["fields"] = merged_fields
        for key in _product_match_keys(merged_fields):
          if key:
            product_index[key] = existing_product
      else:
        unchanged_count += 1

      row_write_ops.append(UpdateOne(
        {"_id": row_id, "import_id": import_id},
        {"$set": {
          "status": "applied",
          "product_id": product_id,
          "applied_at": now,
          "applied_fields": sorted(changed_fields.keys()),
          "updated_at": now,
        }},
      ))
      continue

    product_id = uuid.uuid4().hex
    new_fields = _product_fields_from_import_fields(product_staged_fields, product_id, selected_columns=selected_columns)
    if not new_fields:
      unchanged_count += 1
      continue

    product_doc = {
      "_id": product_id,
      "fields": new_fields,
      "created_at": now,
      "updated_at": now,
      "source_import_id": import_id,
      "source_import_row_id": row_id,
    }
    product_write_ops.append(InsertOne(product_doc))
    created_count += 1
    changed_cells_count += len(new_fields)
    created_product_ids.append(product_id)
    audit_entries = _build_field_change_entries_from_create(new_fields)
    if audit_entries:
      pending_import_audit_logs.append((product_id, audit_entries, "created"))
    for key in _product_match_keys(new_fields):
      if key:
        product_index[key] = product_doc

    row_write_ops.append(UpdateOne(
      {"_id": row_id, "import_id": import_id},
      {"$set": {
        "status": "applied",
        "product_id": product_id,
        "applied_at": now,
        "applied_fields": sorted(new_fields.keys()),
        "updated_at": now,
      }},
    ))

  if product_write_ops:
    await db["products"].bulk_write(product_write_ops, ordered=False)

  if row_write_ops:
    await db[PRODUCT_IMPORT_ROWS_COLLECTION].bulk_write(row_write_ops, ordered=False)

  await _insert_import_apply_field_audit_logs(
    req,
    user=user,
    db=db,
    import_label=import_label,
    pending=pending_import_audit_logs,
  )

  await db[PRODUCT_IMPORT_BATCHES_COLLECTION].update_one(
    {"_id": import_id},
    {
      "$set": {
        "status": "applied",
        "applied_at": now,
        "updated_at": now,
        "apply_summary": {
          "created_count": created_count,
          "updated_count": updated_count,
          "unchanged_count": unchanged_count,
          "changed_cells_count": changed_cells_count,
          "selected_columns": sorted(selected_columns),
          "match": (
            {
              "import_column": import_match_column,
              "product_column": product_match_column,
            }
            if match_config
            else None
          ),
          "apply_row_groups": sorted(apply_row_groups) if match_config else None,
          "skipped_count": skipped_count,
          "rows_processed_by_group": rows_processed_by_group,
          "rows_skipped_by_group": rows_skipped_by_group,
        },
      }
    },
  )

  await log_activity(
    req,
    "PRODUCT_IMPORT_APPLY",
    details=f"Applied product import: {batch.get('filename')}. Created: {created_count}, updated: {updated_count}, unchanged: {unchanged_count}",
    resource_id=import_id,
    user=user,
    db=db,
  )

  return {
    "ok": True,
    "created_count": created_count,
    "updated_count": updated_count,
    "unchanged_count": unchanged_count,
    "skipped_count": skipped_count,
    "changed_cells_count": changed_cells_count,
    "created_product_ids": created_product_ids[:50],
    "updated_product_ids": updated_product_ids[:50],
    "apply_row_groups": sorted(apply_row_groups) if match_config else None,
    "rows_processed_by_group": rows_processed_by_group,
    "rows_skipped_by_group": rows_skipped_by_group,
  }


@api.delete("/admin/products/imports/{import_id}")
async def admin_delete_product_import(
  import_id: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  batch = await db[PRODUCT_IMPORT_BATCHES_COLLECTION].find_one({"_id": import_id})
  if not batch:
    raise HTTPException(status_code=404, detail="IMPORT_NOT_FOUND")

  rows_result = await db[PRODUCT_IMPORT_ROWS_COLLECTION].delete_many({"import_id": import_id})
  await db[PRODUCT_IMPORT_BATCHES_COLLECTION].delete_one({"_id": import_id})
  shutil.rmtree(STORAGE_DIR / "product-imports" / import_id, ignore_errors=True)

  await log_activity(
    req,
    "PRODUCT_IMPORT_DELETE",
    details=f"Deleted product import: {batch.get('filename')}. Rows: {rows_result.deleted_count}",
    resource_id=import_id,
    user=user,
    db=db,
  )

  return {"ok": True, "deleted_rows": rows_result.deleted_count}


@api.get("/admin/backups")
async def admin_list_backups(
  _: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  """List all backup collections in the database."""
  all_collections = await db.list_collection_names()
  backups: list[dict[str, Any]] = []
  for name in sorted(all_collections):
    if "_backup_" not in name:
      continue
    parts = name.split("_backup_", 1)
    original = parts[0]
    timestamp_str = parts[1] if len(parts) > 1 else ""
    count = await db[name].count_documents({})
    backups.append({
      "name": name,
      "original_collection": original,
      "timestamp": timestamp_str,
      "record_count": count,
    })
  return {"backups": backups}


@api.post("/admin/backup/{collection_name}")
async def admin_create_backup(
  collection_name: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  """Create a backup of a collection by duplicating it."""
  if collection_name not in _BACKUPABLE_COLLECTIONS:
    raise HTTPException(status_code=400, detail=f"Collection '{collection_name}' is not backupable. Allowed: {_BACKUPABLE_COLLECTIONS}")

  now = datetime.now(timezone.utc)
  ts = now.strftime("%Y%m%d_%H%M%S")
  backup_name = f"{collection_name}_backup_{ts}"

  # Check if backup already exists
  existing = await db.list_collection_names()
  if backup_name in existing:
    raise HTTPException(status_code=409, detail=f"Backup '{backup_name}' already exists")

  count = 0
  stage_name = f"{backup_name}_staging_{uuid.uuid4().hex[:8]}"
  try:
    batch: list[dict[str, Any]] = []
    async for doc in db[collection_name].find({}):
      batch.append(doc)
      if len(batch) >= 500:
        await db[stage_name].insert_many(batch, ordered=False)
        count += len(batch)
        batch = []
    if batch:
      await db[stage_name].insert_many(batch, ordered=False)
      count += len(batch)

    async for idx in db[collection_name].list_indexes():
      key = idx.get("key")
      if key == {"_id": 1}:
        continue
      keys = list(key.items()) if isinstance(key, dict) else key
      if not keys:
        continue
      opts = {k: v for k, v in idx.items() if k not in {"v", "ns", "key"}}
      await db[stage_name].create_index(keys, **opts)

    await db.client.admin.command(
      "renameCollection",
      f"{db.name}.{stage_name}",
      to=f"{db.name}.{backup_name}",
      dropTarget=False,
    )
  except Exception:
    try:
      await db[stage_name].drop()
    except Exception:
      pass
    raise

  await log_activity(req, "BACKUP_CREATE", details=f"Created backup: {backup_name} ({count} records)", user=user, db=db)

  return {"backup_name": backup_name, "record_count": count, "created_at": now.isoformat()}


@api.post("/admin/restore/{backup_name}")
async def admin_restore_backup(
  backup_name: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  """Restore a backup by replacing the original collection's documents."""
  if "_backup_" not in backup_name:
    raise HTTPException(status_code=400, detail="Invalid backup name format")

  original_name = backup_name.split("_backup_", 1)[0]
  if original_name not in _BACKUPABLE_COLLECTIONS:
    raise HTTPException(status_code=400, detail=f"Cannot restore to '{original_name}'")

  # Check backup exists
  existing = await db.list_collection_names()
  if backup_name not in existing:
    raise HTTPException(status_code=404, detail=f"Backup '{backup_name}' not found")

  backup_count = await db[backup_name].count_documents({})
  if backup_count == 0:
    raise HTTPException(status_code=400, detail="Backup is empty")
  stage_name = f"{original_name}_restore_stage_{uuid.uuid4().hex[:8]}"
  archived_name = f"{original_name}_pre_restore_{uuid.uuid4().hex[:8]}"
  original_renamed = False
  try:
    batch: list[dict[str, Any]] = []
    async for doc in db[backup_name].find({}):
      batch.append(doc)
      if len(batch) >= 500:
        await db[stage_name].insert_many(batch, ordered=False)
        batch = []
    if batch:
      await db[stage_name].insert_many(batch, ordered=False)

    async for idx in db[backup_name].list_indexes():
      key = idx.get("key")
      if key == {"_id": 1}:
        continue
      keys = list(key.items()) if isinstance(key, dict) else key
      if not keys:
        continue
      opts = {k: v for k, v in idx.items() if k not in {"v", "ns", "key"}}
      await db[stage_name].create_index(keys, **opts)

    if original_name in existing:
      await db.client.admin.command(
        "renameCollection",
        f"{db.name}.{original_name}",
        to=f"{db.name}.{archived_name}",
        dropTarget=False,
      )
      original_renamed = True

    await db.client.admin.command(
      "renameCollection",
      f"{db.name}.{stage_name}",
      to=f"{db.name}.{original_name}",
      dropTarget=False,
    )

    if original_renamed:
      await db[archived_name].drop()
  except Exception as exc:
    if original_renamed:
      try:
        names_after = await db.list_collection_names()
        if original_name not in names_after and archived_name in names_after:
          await db.client.admin.command(
            "renameCollection",
            f"{db.name}.{archived_name}",
            to=f"{db.name}.{original_name}",
            dropTarget=False,
          )
      except Exception:
        pass
    try:
      await db[stage_name].drop()
    except Exception:
      pass
    raise HTTPException(status_code=500, detail=f"RESTORE_FAILED: {exc}")

  await log_activity(req, "BACKUP_RESTORE", details=f"Restored '{original_name}' from '{backup_name}' ({backup_count} records)", user=user, db=db)

  return {"ok": True, "restored_collection": original_name, "record_count": backup_count}


@api.delete("/admin/backup/{backup_name}")
async def admin_delete_backup(
  backup_name: str,
  req: FastAPIRequest,
  user: dict[str, Any] = Depends(_require_admin),
  db: Any = Depends(_get_db),
):
  """Delete a backup collection."""
  if "_backup_" not in backup_name:
    raise HTTPException(status_code=400, detail="Invalid backup name format")

  existing = await db.list_collection_names()
  if backup_name not in existing:
    raise HTTPException(status_code=404, detail=f"Backup '{backup_name}' not found")

  await db[backup_name].drop()
  await log_activity(req, "BACKUP_DELETE", details=f"Deleted backup: {backup_name}", user=user, db=db)

  return {"ok": True, "deleted": backup_name}


@api.get("/products/assets")
async def products_assets(
  _: dict[str, Any] = Depends(_get_current_user), 
  limit: int = Query(PRODUCTS_ASSETS_DEFAULT_LIMIT, ge=1, le=PRODUCTS_ASSETS_MAX_LIMIT),
  cursor: str | None = Query(None),
  db=Depends(_get_db)
):
  return await _products_assets_page(db, limit=limit, cursor=cursor)


@api.get("/dam/collection-code")
async def dam_collection_code(collection_name: str, db=Depends(_get_db)):
  name = collection_name.strip()
  if not name:
    raise HTTPException(status_code=400, detail="COLLECTION_NAME_REQUIRED")

  # Search in dam_assets where "Collection Name" matches (case-insensitive)
  # Airtable filter formula was: LOWER({Collection Name})=LOWER("name")
  import re
  query = {"fields.Collection Name": re.compile(f"^{re.escape(name)}$", re.IGNORECASE)}
  
  doc = await db["dam_assets"].find_one(query)
  
  if not doc or not doc.get("fields"):
    return {"collection_name": name, "collection_code": None, "variant_number": None, "price": None}

  fields = doc["fields"]
  
  def _format_field(val):
    if val is None: return None
    if isinstance(val, (int, float)):
      return str(int(val)) if float(val).is_integer() else str(val)
    if isinstance(val, str):
      return val.strip()
    return str(val)

  return {
    "collection_name": name,
    "collection_code": _format_field(fields.get("Collection Code")),
    "variant_number": _format_field(fields.get("Variant Number")),
    "price": _format_field(fields.get("Price")),
  }


def _load_classes() -> list[ClassItem]:
  data = _read_json(CLASSES_PATH, [])
  if not isinstance(data, list):
    return []
  out: list[ClassItem] = []
  for item in data:
    if not isinstance(item, dict):
      continue
    cid = item.get("id")
    name = item.get("name")
    if isinstance(cid, str) and isinstance(name, str):
      out.append({"id": cid, "name": name})
  return out


def _save_classes(items: list[ClassItem]) -> None:
  _safe_write_json(CLASSES_PATH, items)


async def _load_classes_from_db(db: Any) -> list[ClassItem]:
  out: list[ClassItem] = []
  async for doc in db[CLASSES_COLLECTION].find({}, {"_id": 1, "name": 1}).sort("_id", 1):
    cid = doc.get("_id")
    name = doc.get("name")
    if isinstance(cid, str) and isinstance(name, str):
      out.append({"id": cid, "name": name})
  return out


async def _migrate_classes_json_to_mongo_if_needed(db: Any) -> None:
  count = await db[CLASSES_COLLECTION].count_documents({})
  if count > 0:
    return
  file_items = _load_classes()
  if not file_items:
    return
  docs = []
  now = _utc_now_iso()
  for item in file_items:
    docs.append({"_id": item["id"], "name": item["name"], "created_at": now, "updated_at": now})
  try:
    await db[CLASSES_COLLECTION].insert_many(docs, ordered=False)
  except Exception:
    pass


@api.get("/classes")
async def get_classes(_: dict[str, Any] = Depends(_require_operator), db: Any = Depends(_get_db)):
  return await _load_classes_from_db(db)


class _CreateClassBody(TypedDict):
  id: str
  name: str


@api.post("/classes")
async def create_class(req: FastAPIRequest, body: _CreateClassBody, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  cid = body.get("id")
  name = body.get("name")
  if not isinstance(cid, str) or not isinstance(name, str):
    raise HTTPException(status_code=400, detail="INVALID_BODY")
  _validate_class_id(cid)

  exists = await db[CLASSES_COLLECTION].find_one({"_id": cid})
  if exists is not None:
    raise HTTPException(status_code=409, detail="CLASS_EXISTS")

  await db[CLASSES_COLLECTION].insert_one({"_id": cid, "name": name, "created_at": _utc_now_iso(), "updated_at": _utc_now_iso()})
  await log_activity(req, "CLASS_CREATE", details=f"Created class: {name} ({cid})", resource_id=cid, user=user, db=db)
  return {"created": True}


class _RenameClassBody(TypedDict):
  name: str


@api.put("/classes/{class_id}")
async def rename_class(class_id: str, req: FastAPIRequest, body: _RenameClassBody, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  name = body.get("name")
  if not isinstance(name, str) or not name:
    raise HTTPException(status_code=400, detail="INVALID_BODY")

  res = await db[CLASSES_COLLECTION].update_one({"_id": class_id}, {"$set": {"name": name, "updated_at": _utc_now_iso()}})
  if res.matched_count == 0:
    raise HTTPException(status_code=404, detail="CLASS_NOT_FOUND")

  await log_activity(req, "CLASS_RENAME", details=f"Renamed class {class_id} to: {name}", resource_id=class_id, user=user, db=db)
  return {"updated": True}


@api.delete("/classes/{class_id}")
async def delete_class(class_id: str, req: FastAPIRequest, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  res = await db[CLASSES_COLLECTION].delete_one({"_id": class_id})
  if res.deleted_count == 0:
    raise HTTPException(status_code=404, detail="CLASS_NOT_FOUND")
  await log_activity(req, "CLASS_DELETE", details=f"Deleted class: {class_id}", resource_id=class_id, user=user, db=db)
  return {"deleted": True}


def _load_queue() -> list[QueueItem]:
  data = _read_json(QUEUE_PATH, [])
  if not isinstance(data, list):
    return []
  out: list[QueueItem] = []
  for item in data:
    if not isinstance(item, dict):
      continue
    item_id = item.get("item_id")
    filename = item.get("filename")
    status = item.get("status")
    created_at = item.get("created_at")
    annotation = item.get("annotation")

    if not isinstance(item_id, str) or not isinstance(filename, str):
      continue
    if status not in ("pending", "labeled"):
      status = "pending"

    qi: QueueItem = {
      "item_id": item_id,
      "filename": filename,
      "status": status,
      "created_at": created_at if isinstance(created_at, str) else _utc_now_iso(),
    }
    if isinstance(annotation, dict):
      qi["annotation"] = annotation  # type: ignore[assignment]
    out.append(qi)

  return out


def _save_queue(items: list[QueueItem]) -> None:
  _safe_write_json(QUEUE_PATH, items)


def _queue_doc_to_item(doc: dict[str, Any]) -> QueueItem | None:
  item_id = doc.get("_id")
  filename = doc.get("filename")
  status = doc.get("status")
  created_at = doc.get("created_at")
  annotation = doc.get("annotation")

  if not isinstance(item_id, str) or not isinstance(filename, str):
    return None
  if status not in ("pending", "labeled"):
    status = "pending"

  qi: QueueItem = {
    "item_id": item_id,
    "filename": filename,
    "status": status,
    "created_at": created_at if isinstance(created_at, str) else _utc_now_iso(),
  }
  if isinstance(annotation, dict):
    qi["annotation"] = annotation  # type: ignore[assignment]
  return qi


async def _load_queue_from_db(db: Any) -> list[QueueItem]:
  out: list[QueueItem] = []
  async for doc in db[QUEUE_COLLECTION].find({}).sort("created_at", -1):
    qi = _queue_doc_to_item(doc)
    if qi is not None:
      out.append(qi)
  return out


async def _migrate_queue_json_to_mongo_if_needed(db: Any) -> None:
  count = await db[QUEUE_COLLECTION].count_documents({})
  if count > 0:
    return
  file_items = _load_queue()
  if not file_items:
    return
  docs: list[dict[str, Any]] = []
  for item in file_items:
    doc: dict[str, Any] = {
      "_id": item["item_id"],
      "filename": item["filename"],
      "status": item.get("status", "pending"),
      "created_at": item.get("created_at", _utc_now_iso()),
    }
    if isinstance(item.get("annotation"), dict):
      doc["annotation"] = item["annotation"]
    docs.append(doc)
  try:
    await db[QUEUE_COLLECTION].insert_many(docs, ordered=False)
  except Exception:
    pass


def _unlink_if_exists(p: Path) -> None:
  try:
    if p.exists():
      p.unlink()
  except Exception:
    return


@api.post("/uploads")
async def upload_image(req: FastAPIRequest, file: UploadFile = File(...), user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  if not file.content_type or not file.content_type.startswith("image/"):
    raise HTTPException(status_code=400, detail="INVALID_IMAGE")

  raw = await file.read()
  try:
    img = Image.open(io.BytesIO(raw)).convert("RGB")
  except Exception:
    raise HTTPException(status_code=400, detail="INVALID_IMAGE")

  item_id = str(uuid.uuid4())
  filename = f"{item_id}.jpg"
  path = UPLOADS_DIR / filename
  img.save(path, format="JPEG", quality=92)

  await db[QUEUE_COLLECTION].insert_one(
    {
      "_id": item_id,
      "filename": filename,
      "status": "pending",
      "created_at": _utc_now_iso(),
    }
  )

  await log_activity(req, "IMAGE_UPLOAD", details=f"Uploaded image: {filename}", resource_id=item_id, user=user, db=db)

  return {
    "item_id": item_id,
    "image_url": f"/files/uploads/{filename}",
  }


@api.get("/queue")
async def get_queue(_: dict[str, Any] = Depends(_require_operator), db: Any = Depends(_get_db)):
  items = await _load_queue_from_db(db)
  for item in items:
    item["image_url"] = f"/files/uploads/{item['filename']}"
  return items


@api.get("/queue/{item_id}")
async def get_queue_item(item_id: str, _: dict[str, Any] = Depends(_require_operator), db: Any = Depends(_get_db)):
  doc = await db[QUEUE_COLLECTION].find_one({"_id": item_id})
  item = _queue_doc_to_item(doc) if isinstance(doc, dict) else None
  if item is not None:
    item["image_url"] = f"/files/uploads/{item['filename']}"
    return item
  raise HTTPException(status_code=404, detail="ITEM_NOT_FOUND")


@api.delete("/queue/{item_id}")
async def delete_queue_item(item_id: str, req: FastAPIRequest, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  target = await db[QUEUE_COLLECTION].find_one({"_id": item_id})
  if target is None:
    raise HTTPException(status_code=404, detail="ITEM_NOT_FOUND")
  await db[QUEUE_COLLECTION].delete_one({"_id": item_id})

  filename = target.get("filename")
  if isinstance(filename, str) and filename:
    _unlink_if_exists(UPLOADS_DIR / filename)

    ds_dir = _dataset_dir()
    stem = Path(filename).stem
    for split in ("train", "val"):
      _unlink_if_exists(ds_dir / "images" / split / filename)
      _unlink_if_exists(ds_dir / "labels" / split / f"{stem}.txt")

  await log_activity(req, "IMAGE_DELETE", details=f"Deleted image: {filename}", resource_id=item_id, user=user, db=db)

  return {"deleted": True}


@api.post("/queue/{item_id}/delete")
async def delete_queue_item_post(item_id: str, req: FastAPIRequest, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  return await delete_queue_item(item_id, req, user, db)


class _SaveAnnotationBody(TypedDict):
  class_id: str
  bbox: NormalizedBBox


def _validate_bbox(b: NormalizedBBox) -> None:
  for k in ("x", "y", "w", "h"):
    v = b.get(k)
    if not isinstance(v, (float, int)):
      raise HTTPException(status_code=400, detail="INVALID_BBOX")
    if float(v) < 0 or float(v) > 1:
      raise HTTPException(status_code=400, detail="INVALID_BBOX")
  if float(b["w"]) <= 0 or float(b["h"]) <= 0:
    raise HTTPException(status_code=400, detail="INVALID_BBOX")


@api.post("/queue/{item_id}/annotation")
async def save_annotation(item_id: str, req: FastAPIRequest, body: _SaveAnnotationBody, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  class_id = body.get("class_id")
  bbox = body.get("bbox")
  if not isinstance(class_id, str) or not isinstance(bbox, dict):
    raise HTTPException(status_code=400, detail="INVALID_BODY")

  _validate_bbox(bbox)  # type: ignore[arg-type]

  classes = await _load_classes_from_db(db)
  if not any(c["id"] == class_id for c in classes):
    raise HTTPException(status_code=400, detail="UNKNOWN_CLASS")

  res = await db[QUEUE_COLLECTION].update_one(
    {"_id": item_id},
    {"$set": {"annotation": {"class_id": class_id, "bbox": bbox}, "status": "labeled"}},
  )
  if res.matched_count == 0:
    raise HTTPException(status_code=404, detail="ITEM_NOT_FOUND")
  return {"saved": True}


def _dataset_dir() -> Path:
  return DATASETS_DIR / "lorenzo_v1"


@api.post("/export")
async def export_dataset(req: FastAPIRequest, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  ds_dir = _dataset_dir()
  images_train = ds_dir / "images" / "train"
  images_val = ds_dir / "images" / "val"
  labels_train = ds_dir / "labels" / "train"
  labels_val = ds_dir / "labels" / "val"

  for d in (images_train, images_val, labels_train, labels_val):
    d.mkdir(parents=True, exist_ok=True)

  classes = await _load_classes_from_db(db)
  class_to_index = {c["id"]: i for i, c in enumerate(classes)}
  if not classes:
    raise HTTPException(status_code=400, detail="NO_CLASSES")

  queue = await _load_queue_from_db(db)
  labeled = [q for q in queue if q.get("status") == "labeled" and isinstance(q.get("annotation"), dict)]
  if not labeled:
    raise HTTPException(status_code=400, detail="NO_LABELED_ITEMS")

  rng = random.Random(42)
  ids = [q["item_id"] for q in labeled if isinstance(q.get("item_id"), str)]
  ids_sorted = sorted(ids)
  rng.shuffle(ids_sorted)

  split = max(1, int(len(ids_sorted) * 0.8))
  train_ids = set(ids_sorted[:split])

  for item in labeled:
    item_id = item["item_id"]
    filename = item["filename"]
    ann = item.get("annotation")
    if not isinstance(ann, dict):
      continue

    class_id = ann.get("class_id")
    bbox = ann.get("bbox")
    if not isinstance(class_id, str) or not isinstance(bbox, dict):
      continue

    class_index = class_to_index.get(class_id)
    if class_index is None:
      continue

    src_img = UPLOADS_DIR / filename
    if not src_img.exists():
      continue

    is_train = item_id in train_ids
    dst_img_dir = images_train if is_train else images_val
    dst_label_dir = labels_train if is_train else labels_val

    dst_img = dst_img_dir / filename
    shutil.copyfile(src_img, dst_img)

    x = float(bbox["x"])
    y = float(bbox["y"])
    w = float(bbox["w"])
    h = float(bbox["h"])

    xc = x + w / 2
    yc = y + h / 2

    label_line = f"{class_index} {xc:.6f} {yc:.6f} {w:.6f} {h:.6f}\n"
    (dst_label_dir / f"{Path(filename).stem}.txt").write_text(label_line, encoding="utf-8")

  data_yaml = {
    "path": str(ds_dir),
    "train": "images/train",
    "val": "images/val",
    "names": [c["name"] for c in classes],
  }

  (ds_dir / "data.yaml").write_text(yaml.safe_dump(data_yaml, sort_keys=False), encoding="utf-8")

  await log_activity(req, "DATASET_EXPORT", details=f"Exported dataset. Items: {len(labeled)}", user=user, db=db)
  return {
    "exported": True,
    "dataset": "lorenzo_v1",
    "path": str(ds_dir),
    "labeled_count": len(labeled),
  }


_jobs_lock = threading.Lock()
_jobs: dict[str, subprocess.Popen[bytes]] = {}


def _job_dir(job_id: str) -> Path:
  return RUNS_DIR / job_id


def _write_job_meta(job_id: str, meta: dict[str, Any]) -> None:
  _safe_write_json(_job_dir(job_id) / "job.json", meta)


def _read_job_meta(job_id: str) -> dict[str, Any] | None:
  p = _job_dir(job_id) / "job.json"
  data = _read_json(p, None)
  return data if isinstance(data, dict) else None


def _tail_lines(path: Path, n: int) -> list[str]:
  if not path.exists():
    return []
  text = path.read_text(encoding="utf-8", errors="ignore")
  lines = text.splitlines()
  return lines[-n:]


def _monitor_job(job_id: str, proc: subprocess.Popen[bytes], log_path: Path, run_dir: Path) -> None:
  status: Literal["running", "finished", "failed"] = "running"
  exit_code: int | None = None

  try:
    exit_code = proc.wait()
    status = "finished" if exit_code == 0 else "failed"
  except Exception:
    status = "failed"

  best_pt = run_dir / "weights" / "best.pt"
  meta = _read_job_meta(job_id) or {}
  meta["status"] = status
  meta["exit_code"] = exit_code
  meta["finished_at"] = _utc_now_iso()
  meta["best_pt"] = str(best_pt) if best_pt.exists() else None
  _write_job_meta(job_id, meta)

  with _jobs_lock:
    _jobs.pop(job_id, None)


class _TrainBody(TypedDict, total=False):
  epochs: int
  batch: int
  imgsz: int


@api.post("/train")
async def start_train(req: FastAPIRequest, body: _TrainBody, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  ds_dir = _dataset_dir()
  data_yaml = ds_dir / "data.yaml"
  if not data_yaml.exists():
    raise HTTPException(status_code=400, detail="DATASET_NOT_EXPORTED")

  epochs = int(body.get("epochs", 50))
  batch = int(body.get("batch", 8))
  imgsz = int(body.get("imgsz", 640))

  if epochs <= 0 or batch <= 0 or imgsz <= 0:
    raise HTTPException(status_code=400, detail="INVALID_PARAMS")

  job_id = str(uuid.uuid4())
  run_dir = _job_dir(job_id)
  run_dir.mkdir(parents=True, exist_ok=True)

  log_path = run_dir / "train.log"

  cmd = [
    "yolo",
    "train",
    "model=yolov8n.pt",
    f"data={str(data_yaml)}",
    f"imgsz={imgsz}",
    f"epochs={epochs}",
    f"batch={batch}",
    f"project={str(run_dir)}",
    "exist_ok=True",
  ]

  with log_path.open("wb") as log_file:
    try:
      proc = subprocess.Popen(
        cmd,
        stdout=log_file,
        stderr=subprocess.STDOUT,
        cwd=str(BASE_DIR),
      )
    except FileNotFoundError:
      raise HTTPException(status_code=500, detail="ULTRALYTICS_NOT_INSTALLED")

  meta = {
    "job_id": job_id,
    "status": "running",
    "started_at": _utc_now_iso(),
    "params": {"epochs": epochs, "batch": batch, "imgsz": imgsz},
    "log_path": str(log_path),
    "run_dir": str(run_dir),
  }
  _write_job_meta(job_id, meta)

  with _jobs_lock:
    _jobs[job_id] = proc

  t = threading.Thread(target=_monitor_job, args=(job_id, proc, log_path, run_dir), daemon=True)
  t.start()

  await log_activity(req, "TRAIN_START", details=f"Started training. Epochs: {epochs}, Batch: {batch}", user=user, db=db)
  return {"job_id": job_id}


def _read_metrics(run_dir: Path) -> dict[str, Any] | None:
  results_csv = run_dir / "results.csv"
  if not results_csv.exists():
    return None

  try:
    rows = results_csv.read_text(encoding="utf-8", errors="ignore").splitlines()
    if len(rows) < 2:
      return None
    header = [h.strip() for h in rows[0].split(",")]
    last = [v.strip() for v in rows[-1].split(",")]
    if len(header) != len(last):
      return None
    d = dict(zip(header, last))

    keys = [
      "metrics/precision(B)",
      "metrics/recall(B)",
      "metrics/mAP50(B)",
      "metrics/mAP50-95(B)",
    ]
    out: dict[str, Any] = {}
    for k in keys:
      if k in d:
        try:
          out[k] = float(d[k])
        except Exception:
          out[k] = d[k]
    return out or None
  except Exception:
    return None


@api.get("/train/{job_id}")
async def get_train_status(job_id: str, lines: int = 120, _: dict[str, Any] = Depends(_require_operator)):
  meta = _read_job_meta(job_id)
  if meta is None:
    raise HTTPException(status_code=404, detail="JOB_NOT_FOUND")

  log_path = Path(str(meta.get("log_path", "")))
  tail = _tail_lines(log_path, max(1, min(int(lines), 500)))

  run_dir = Path(str(meta.get("run_dir", "")))
  metrics = _read_metrics(run_dir)

  return {
    "job_id": job_id,
    "status": meta.get("status"),
    "best_pt": meta.get("best_pt"),
    "metrics": metrics,
    "log": tail,
  }


@api.post("/train/{job_id}/publish")
async def publish(job_id: str, req: FastAPIRequest, user: dict[str, Any] = Depends(_get_current_user), db: Any = Depends(_get_db)):
  _require_role(user, {"admin", "sales"})
  meta = _read_job_meta(job_id)
  if meta is None:
    raise HTTPException(status_code=404, detail="JOB_NOT_FOUND")

  best_pt = meta.get("best_pt")
  if not isinstance(best_pt, str) or not best_pt:
    raise HTTPException(status_code=400, detail="BEST_PT_NOT_AVAILABLE")

  src = Path(best_pt)
  if not src.exists():
    raise HTTPException(status_code=400, detail="BEST_PT_NOT_AVAILABLE")

  repo_root = BASE_DIR.parent.parent
  dst = repo_root / "backend" / "models" / "best.pt"
  dst.parent.mkdir(parents=True, exist_ok=True)
  shutil.copyfile(src, dst)

  await log_activity(req, "TRAIN_PUBLISH", details=f"Published model weights for job: {job_id}", resource_id=job_id, user=user, db=db)
  return {"published": True, "path": str(dst)}
